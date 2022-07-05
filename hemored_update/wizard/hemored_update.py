# -*- coding: utf-8 -*-
import base64
import os
import tempfile
from odoo import fields, api, models
from odoo.exceptions import ValidationError
from openpyxl import load_workbook

from ..models import constants


class UpdateExcel(models.TransientModel):
    _name = 'hemored_update.updateexcel'

    ficha_estadistica = fields.Binary(
        string='Ficha Estadística',
        attachment=True,

    )
    txt_ficha_estadistica = fields.Char(
        string='Nombre Ficha Estadística'
    )

    @api.multi
    def is_int(self, valor):
        try:
            if valor is None:
                return False
            else:
                int(valor)
                return True
        except ValueError:
            return False

    @api.multi
    def cargar_excel_ficha(self):
        ficha = self.env.context.get('active_id')
        ficha_id = self.env['hemored.ficha_estadistica'].browse(ficha)
        if ficha_id.state != constants.VALIDADO:
            data = base64.decodestring(self.ficha_estadistica)
            fobj = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            fname = fobj.name
            fobj.write(data)
            fobj.close()
            book = load_workbook(fname)
            sheets = book.sheetnames
            ws = book[sheets[0]]

            if ficha_id.banco_id.renipress_id.codigo_eess == ws['J3'].value and (ficha_id.periodo_id.name == ws['I10'].value or ficha_id.anho_id.name == ws['I10'].value) and fields.Date.today() <= ficha_id.fin_fecha_registro:
                if ficha_id.banco_id.tipo_banco == '2':
                    cantidad_mvih_s_o_n = constants.SI if ws['C69'].value == 'X' or ws['C69'].value == 'x' else constants.NO
                    cantidad_mhbsag_s_o_n = constants.SI if ws['D69'].value == 'X' or ws['D69'].value == 'x' else constants.NO
                    cantidad_mhepc_s_o_n = constants.SI if ws['E69'].value == 'X' or ws['E69'].value == 'x' else constants.NO
                    cantidad_mantihbc_s_o_n = constants.SI if ws['F69'].value == 'X' or ws['F69'].value == 'x' else constants.NO
                    cantidad_mhtlv_s_o_n = constants.SI if ws['G69'].value == 'X' or ws['G69'].value == 'x' else constants.NO
                    cantidad_msifilis_s_o_n = constants.SI if ws['H69'].value == 'X' or ws['H69'].value == 'x' else constants.NO
                    cantidad_mchagas_s_o_n = constants.SI if ws['I69'].value == 'X' or ws['I69'].value == 'x' else constants.NO
                    cantidad_motros_s_o_n = constants.SI if ws['J69'].value == 'X' or ws['J69'].value == 'x' else constants.NO

                    donantes_line = self.env['hemored.ficha_total_donante.line'].search([('ficha_id', '=', ficha_id.id)])
                    total_donantes_line = 0
                    for line in donantes_line:
                        line.write({
                            'cantidad': int(ws[constants.TIPO_DONANTE[line.tipo_donante]].value) if self.is_int(ws[constants.TIPO_DONANTE[line.tipo_donante]].value) else 0,
                        })
                        total_donantes_line += int(ws[constants.TIPO_DONANTE[line.tipo_donante]].value) if self.is_int(ws[constants.TIPO_DONANTE[line.tipo_donante]].value) else 0

                    postulante_line = self.env['hemored.ficha_total_postulante_diferido_excluido.line'].search([('ficha_id', '=', ficha_id.id)])
                    total_postulante_line = 0
                    for line in postulante_line:
                        line.write({
                            'cantidad': int(ws[constants.TIPO_POSTULANTE[line.motivo_diferido_excluido]].value) if self.is_int(ws[constants.TIPO_POSTULANTE[line.motivo_diferido_excluido]].value) else 0,
                        })
                        total_postulante_line += int(ws[constants.TIPO_POSTULANTE[line.motivo_diferido_excluido]].value) if self.is_int(ws[constants.TIPO_POSTULANTE[line.motivo_diferido_excluido]].value) else 0

                    donacion_grupo_edad_line = self.env['hemored.ficha_donacion_segun_grupo_edad_donante.line'].search([('ficha_id', '=', ficha_id.id)])
                    total_donacion_grupo_edad_line = 0
                    for line in donacion_grupo_edad_line:
                        line.write({
                            'cantidad': int(ws[constants.GRUPO_EDAD[line.grupo_edad_donante]].value) if self.is_int(ws[constants.GRUPO_EDAD[line.grupo_edad_donante]].value) else 0,
                        })
                        total_donacion_grupo_edad_line += int(ws[constants.GRUPO_EDAD[line.grupo_edad_donante]].value) if self.is_int(ws[constants.GRUPO_EDAD[line.grupo_edad_donante]].value) else 0

                    sangre_total_line = self.env['hemored.ficha_donacion_sangre_total.line'].search([('ficha_id', '=', ficha_id.id)])
                    total_sangre_total_line = 0
                    for line in sangre_total_line:
                        line.write({
                            'cantidad': int(ws[constants.SANGRE_TOTAL[line.tipo_donante]].value) if self.is_int(ws[constants.SANGRE_TOTAL[line.tipo_donante]].value) else 0,
                        })
                        total_sangre_total_line += int(ws[constants.SANGRE_TOTAL[line.tipo_donante]].value) if self.is_int(ws[constants.SANGRE_TOTAL[line.tipo_donante]].value) else 0

                    donacion_aferesis_line = self.env['hemored.ficha_donacion_aferesis.line'].search([('ficha_id', '=', ficha_id.id)])
                    total_donacion_aferesis_line = 0
                    for line in donacion_aferesis_line:
                        line.write({
                            'cantidad': int(ws[constants.DONACION_AFERESIS[line.tipo_donante]].value) if self.is_int(ws[constants.DONACION_AFERESIS[line.tipo_donante]].value) else 0,
                        })
                        total_donacion_aferesis_line += int(ws[constants.DONACION_AFERESIS[line.tipo_donante]].value) if self.is_int(ws[constants.DONACION_AFERESIS[line.tipo_donante]].value) else 0

                    produccion_unidad_sangre_line = self.env['hemored.ficha_produccion_unidad_sangre.line'].search([('ficha_id', '=', ficha_id.id)])
                    total_produccion_unidad_sangre_line = 0
                    for line in produccion_unidad_sangre_line:
                        line.write({
                            'cantidad_tipo_unidad_sangre': int(ws[constants.TIPO_UNIDAD_SANGRE[line.tipo_unidad_sangre]].value) if self.is_int(ws[constants.TIPO_UNIDAD_SANGRE[line.tipo_unidad_sangre]].value) else 0,
                        })
                        total_produccion_unidad_sangre_line += int(ws[constants.TIPO_UNIDAD_SANGRE[line.tipo_unidad_sangre]].value) if self.is_int(ws[constants.TIPO_UNIDAD_SANGRE[line.tipo_unidad_sangre]].value) else 0

                    produccion_hemocomponente_line = self.env['hemored.ficha_produccion_hemocomponente.line'].search([('ficha_id', '=', ficha_id.id)])
                    total_produccion_hemocomponente_line = 0
                    for line in produccion_hemocomponente_line:
                        line.write({
                            'cantidad_hemocomponente': int(ws[constants.PRODUCCION_HEMOCOMPONENTES[line.hemocomponente]].value) if self.is_int(ws[constants.PRODUCCION_HEMOCOMPONENTES[line.hemocomponente]].value) else 0,
                        })
                        total_produccion_hemocomponente_line += int(ws[constants.PRODUCCION_HEMOCOMPONENTES[line.hemocomponente]].value) if self.is_int(ws[constants.PRODUCCION_HEMOCOMPONENTES[line.hemocomponente]].value) else 0

                    ficha_id.write({
                        'cantidad_pe_x_ca': int(ws['C15'].value) if self.is_int(ws['C15'].value) else 0,
                        'cantidad_pe_x_cvim': int(ws['C16'].value) if self.is_int(ws['C16'].value) else 0,
                        'cantidad_pe_x_cvem': int(ws['C17'].value) if self.is_int(ws['C17'].value) else 0,
                        'cantidad_pe_x_cr': int(ws['C18'].value) if self.is_int(ws['C18'].value) else 0,
                        'cantidad_pe_x_cpr': int(ws['C19'].value) if self.is_int(ws['C19'].value) else 0,

                        'cantidad_pd_x_ca': int(ws['D15'].value) if self.is_int(ws['D15'].value) else 0,
                        'cantidad_pd_x_cvim': int(ws['D16'].value) if self.is_int(ws['D16'].value) else 0,
                        'cantidad_pd_x_cvem': int(ws['D17'].value) if self.is_int(ws['D17'].value) else 0,
                        'cantidad_pd_x_cr': int(ws['D18'].value) if self.is_int(ws['D18'].value) else 0,
                        'cantidad_pd_x_cpr': int(ws['D19'].value) if self.is_int(ws['D19'].value) else 0,

                        'cantidad_padi_x_ca': int(ws['E15'].value) if self.is_int(ws['E15'].value) else 0,
                        'cantidad_padi_x_cvim': int(ws['E16'].value) if self.is_int(ws['E16'].value) else 0,
                        'cantidad_padi_x_cvem': int(ws['E17'].value) if self.is_int(ws['E17'].value) else 0,
                        'cantidad_padi_x_cr': int(ws['E18'].value) if self.is_int(ws['E18'].value) else 0,
                        'cantidad_padi_x_cpr': int(ws['E19'].value) if self.is_int(ws['E19'].value) else 0,

                        'cantidad_padc_x_ca': int(ws['F15'].value) if self.is_int(ws['F15'].value) else 0,
                        'cantidad_padc_x_cvim': int(ws['F16'].value) if self.is_int(ws['F16'].value) else 0,
                        'cantidad_padc_x_cvem': int(ws['F17'].value) if self.is_int(ws['F17'].value) else 0,
                        'cantidad_padc_x_cr': int(ws['F18'].value) if self.is_int(ws['F18'].value) else 0,
                        'cantidad_padc_x_cpr': int(ws['F19'].value) if self.is_int(ws['F19'].value) else 0,

                        'cantidad_donacion_h': int(ws['B31'].value) if self.is_int(ws['B31'].value) else 0,
                        'cantidad_donacion_m': int(ws['C31'].value) if self.is_int(ws['C31'].value) else 0,

                        'cantidad_mvih_clt_cv': int(ws['C43'].value) if self.is_int(ws['C43'].value) else 0,
                        'cantidad_mvih_clt_cr': int(ws['C44'].value) if self.is_int(ws['C44'].value) else 0,
                        'cantidad_mvih_clt_cpr': int(ws['C45'].value) if self.is_int(ws['C45'].value) else 0,
                        'cantidad_mvih_clt_ca': int(ws['C46'].value) if self.is_int(ws['C46'].value) else 0,

                        'cantidad_mvih_clr_cv': int(ws['D43'].value) if self.is_int(ws['D43'].value) else 0,
                        'cantidad_mvih_clr_cr': int(ws['D44'].value) if self.is_int(ws['D44'].value) else 0,
                        'cantidad_mvih_clr_cpr': int(ws['D45'].value) if self.is_int(ws['D45'].value) else 0,
                        'cantidad_mvih_clr_ca': int(ws['D46'].value) if self.is_int(ws['D46'].value) else 0,

                        'cantidad_mvih_clzg_cv': int(ws['E43'].value) if self.is_int(ws['E43'].value) else 0,
                        'cantidad_mvih_clzg_cr': int(ws['E44'].value) if self.is_int(ws['E44'].value) else 0,
                        'cantidad_mvih_clzg_cpr': int(ws['E45'].value) if self.is_int(ws['E45'].value) else 0,
                        'cantidad_mvih_clzg_ca': int(ws['E46'].value) if self.is_int(ws['E46'].value) else 0,

                        'cantidad_mhbsag_clt_cv': int(ws['F43'].value) if self.is_int(ws['F43'].value) else 0,
                        'cantidad_mhbsag_clt_cr': int(ws['F44'].value) if self.is_int(ws['F44'].value) else 0,
                        'cantidad_mhbsag_clt_cpr': int(ws['F45'].value) if self.is_int(ws['F45'].value) else 0,
                        'cantidad_mhbsag_clt_ca': int(ws['F46'].value) if self.is_int(ws['F46'].value) else 0,
                        'cantidad_mhbsag_clr_cv': int(ws['G43'].value) if self.is_int(ws['G43'].value) else 0,
                        'cantidad_mhbsag_clr_cr': int(ws['G44'].value) if self.is_int(ws['G44'].value) else 0,
                        'cantidad_mhbsag_clr_cpr': int(ws['G45'].value) if self.is_int(ws['G45'].value) else 0,
                        'cantidad_mhbsag_clr_ca': int(ws['G46'].value) if self.is_int(ws['G46'].value) else 0,

                        'cantidad_mhbsag_clzg_cv': int(ws['H43'].value) if self.is_int(ws['H43'].value) else 0,
                        'cantidad_mhbsag_clzg_cr': int(ws['H44'].value) if self.is_int(ws['H44'].value) else 0,
                        'cantidad_mhbsag_clzg_cpr': int(ws['H45'].value) if self.is_int(ws['H45'].value) else 0,
                        'cantidad_mhbsag_clzg_ca': int(ws['H46'].value) if self.is_int(ws['H46'].value) else 0,

                        'cantidad_mhepc_clt_cv': int(ws['I43'].value) if self.is_int(ws['I43'].value) else 0,
                        'cantidad_mhepc_clt_cr': int(ws['I44'].value) if self.is_int(ws['I44'].value) else 0,
                        'cantidad_mhepc_clt_cpr': int(ws['I45'].value) if self.is_int(ws['I45'].value) else 0,
                        'cantidad_mhepc_clt_ca': int(ws['I46'].value) if self.is_int(ws['I46'].value) else 0,

                        'cantidad_mhepc_clr_cv': int(ws['J43'].value) if self.is_int(ws['J43'].value) else 0,
                        'cantidad_mhepc_clr_cr': int(ws['J44'].value) if self.is_int(ws['J44'].value) else 0,
                        'cantidad_mhepc_clr_cpr': int(ws['J45'].value) if self.is_int(ws['J45'].value) else 0,
                        'cantidad_mhepc_clr_ca': int(ws['J46'].value) if self.is_int(ws['J46'].value) else 0,

                        'cantidad_mhepc_clzg_cv': int(ws['K43'].value) if self.is_int(ws['K43'].value) else 0,
                        'cantidad_mhepc_clzg_cr': int(ws['K44'].value) if self.is_int(ws['K44'].value) else 0,
                        'cantidad_mhepc_clzg_cpr': int(ws['K45'].value) if self.is_int(ws['K45'].value) else 0,
                        'cantidad_mhepc_clzg_ca': int(ws['K46'].value) if self.is_int(ws['K46'].value) else 0,

                        'cantidad_mantihbc_clt_cv': int(ws['L43'].value) if self.is_int(ws['L43'].value) else 0,
                        'cantidad_mantihbc_clt_cr': int(ws['L44'].value) if self.is_int(ws['L44'].value) else 0,
                        'cantidad_mantihbc_clt_cpr': int(ws['L45'].value) if self.is_int(ws['L45'].value) else 0,
                        'cantidad_mantihbc_clt_ca': int(ws['L46'].value) if self.is_int(ws['L46'].value) else 0,

                        'cantidad_mantihbc_clr_cv': int(ws['M43'].value) if self.is_int(ws['M43'].value) else 0,
                        'cantidad_mantihbc_clr_cr': int(ws['M44'].value) if self.is_int(ws['M44'].value) else 0,
                        'cantidad_mantihbc_clr_cpr': int(ws['M45'].value) if self.is_int(ws['M45'].value) else 0,
                        'cantidad_mantihbc_clr_ca': int(ws['M46'].value) if self.is_int(ws['M46'].value) else 0,

                        'cantidad_mantihbc_clzg_cv': int(ws['N43'].value) if self.is_int(ws['N43'].value) else 0,
                        'cantidad_mantihbc_clzg_cr': int(ws['N44'].value) if self.is_int(ws['N44'].value) else 0,
                        'cantidad_mantihbc_clzg_cpr': int(ws['N45'].value) if self.is_int(ws['N45'].value) else 0,
                        'cantidad_mantihbc_clzg_ca': int(ws['N46'].value) if self.is_int(ws['N46'].value) else 0,

                        'cantidad_mhtlv_clt_cv': int(ws['C51'].value) if self.is_int(ws['C51'].value) else 0,
                        'cantidad_mhtlv_clt_cr': int(ws['C52'].value) if self.is_int(ws['C52'].value) else 0,
                        'cantidad_mhtlv_clt_cpr': int(ws['C53'].value) if self.is_int(ws['C53'].value) else 0,
                        'cantidad_mhtlv_clt_ca': int(ws['C54'].value) if self.is_int(ws['C54'].value) else 0,

                        'cantidad_mhtlv_clr_cv': int(ws['D51'].value) if self.is_int(ws['D51'].value) else 0,
                        'cantidad_mhtlv_clr_cr': int(ws['D52'].value) if self.is_int(ws['D52'].value) else 0,
                        'cantidad_mhtlv_clr_cpr': int(ws['D53'].value) if self.is_int(ws['D53'].value) else 0,
                        'cantidad_mhtlv_clr_ca': int(ws['D54'].value) if self.is_int(ws['D54'].value) else 0,

                        'cantidad_mhtlv_clzg_cv': int(ws['E51'].value) if self.is_int(ws['E51'].value) else 0,
                        'cantidad_mhtlv_clzg_cr': int(ws['E52'].value) if self.is_int(ws['E52'].value) else 0,
                        'cantidad_mhtlv_clzg_cpr': int(ws['E53'].value) if self.is_int(ws['E53'].value) else 0,
                        'cantidad_mhtlv_clzg_ca': int(ws['E54'].value) if self.is_int(ws['E54'].value) else 0,

                        'cantidad_msifilis_clt_cv': int(ws['F51'].value) if self.is_int(ws['F51'].value) else 0,
                        'cantidad_msifilis_clt_cr': int(ws['F52'].value) if self.is_int(ws['F52'].value) else 0,
                        'cantidad_msifilis_clt_cpr': int(ws['F53'].value) if self.is_int(ws['F53'].value) else 0,
                        'cantidad_msifilis_clt_ca': int(ws['F54'].value) if self.is_int(ws['F54'].value) else 0,

                        'cantidad_msifilis_clr_cv': int(ws['G51'].value) if self.is_int(ws['G51'].value) else 0,
                        'cantidad_msifilis_clr_cr': int(ws['G52'].value) if self.is_int(ws['G52'].value) else 0,
                        'cantidad_msifilis_clr_cpr': int(ws['G53'].value) if self.is_int(ws['G53'].value) else 0,
                        'cantidad_msifilis_clr_ca': int(ws['G54'].value) if self.is_int(ws['G54'].value) else 0,

                        'cantidad_msifilis_clzg_cv': int(ws['H51'].value) if self.is_int(ws['H51'].value) else 0,
                        'cantidad_msifilis_clzg_cr': int(ws['H52'].value) if self.is_int(ws['H52'].value) else 0,
                        'cantidad_msifilis_clzg_cpr': int(ws['H53'].value) if self.is_int(ws['H53'].value) else 0,
                        'cantidad_msifilis_clzg_ca': int(ws['H54'].value) if self.is_int(ws['H54'].value) else 0,

                        'cantidad_mchagas_clt_cv': int(ws['I51'].value) if self.is_int(ws['I51'].value) else 0,
                        'cantidad_mchagas_clt_cr': int(ws['I52'].value) if self.is_int(ws['I52'].value) else 0,
                        'cantidad_mchagas_clt_cpr': int(ws['I53'].value) if self.is_int(ws['I53'].value) else 0,
                        'cantidad_mchagas_clt_ca': int(ws['I54'].value) if self.is_int(ws['I54'].value) else 0,

                        'cantidad_mchagas_clr_cv': int(ws['J51'].value) if self.is_int(ws['J51'].value) else 0,
                        'cantidad_mchagas_clr_cr': int(ws['J52'].value) if self.is_int(ws['J52'].value) else 0,
                        'cantidad_mchagas_clr_cpr': int(ws['J53'].value) if self.is_int(ws['J53'].value) else 0,
                        'cantidad_mchagas_clr_ca': int(ws['J54'].value) if self.is_int(ws['J54'].value) else 0,

                        'cantidad_mchagas_clzg_cv': int(ws['K51'].value) if self.is_int(ws['K51'].value) else 0,
                        'cantidad_mchagas_clzg_cr': int(ws['K52'].value) if self.is_int(ws['K52'].value) else 0,
                        'cantidad_mchagas_clzg_cpr': int(ws['K53'].value) if self.is_int(ws['K53'].value) else 0,
                        'cantidad_mchagas_clzg_ca': int(ws['K54'].value) if self.is_int(ws['K54'].value) else 0,

                        'cantidad_motros_clt_cv': int(ws['L51'].value) if self.is_int(ws['L51'].value) else 0,
                        'cantidad_motros_clt_cr': int(ws['L52'].value) if self.is_int(ws['L52'].value) else 0,
                        'cantidad_motros_clt_cpr': int(ws['L53'].value) if self.is_int(ws['L53'].value) else 0,
                        'cantidad_motros_clt_ca': int(ws['L54'].value) if self.is_int(ws['L54'].value) else 0,

                        'cantidad_motros_clr_cv': int(ws['M51'].value) if self.is_int(ws['M51'].value) else 0,
                        'cantidad_motros_clr_cr': int(ws['M52'].value) if self.is_int(ws['M52'].value) else 0,
                        'cantidad_motros_clr_cpr': int(ws['M53'].value) if self.is_int(ws['M53'].value) else 0,
                        'cantidad_motros_clr_ca': int(ws['M54'].value) if self.is_int(ws['M54'].value) else 0,

                        'cantidad_motros_clzg_cv': int(ws['N51'].value) if self.is_int(ws['N51'].value) else 0,
                        'cantidad_motros_clzg_cr': int(ws['N52'].value) if self.is_int(ws['N52'].value) else 0,
                        'cantidad_motros_clzg_cpr': int(ws['N53'].value) if self.is_int(ws['N53'].value) else 0,
                        'cantidad_motros_clzg_ca': int(ws['N54'].value) if self.is_int(ws['N54'].value) else 0,

                        'cantidad_unr_cv': int(ws['C60'].value) if self.is_int(ws['C60'].value) else 0,
                        'cantidad_unr_cr': int(ws['C61'].value) if self.is_int(ws['C61'].value) else 0,
                        'cantidad_unr_cpr': int(ws['C62'].value) if self.is_int(ws['C62'].value) else 0,
                        'cantidad_unr_ca': int(ws['C63'].value) if self.is_int(ws['C63'].value) else 0,

                        'cantidad_ur_cv': int(ws['D60'].value) if self.is_int(ws['D60'].value) else 0,
                        'cantidad_ur_cr': int(ws['D61'].value) if self.is_int(ws['D61'].value) else 0,
                        'cantidad_ur_cpr': int(ws['D62'].value) if self.is_int(ws['D62'].value) else 0,
                        'cantidad_ur_ca': int(ws['D63'].value) if self.is_int(ws['D63'].value) else 0,

                        'cantidad_uzg_cv': int(ws['E60'].value) if self.is_int(ws['E60'].value) else 0,
                        'cantidad_uzg_cr': int(ws['E61'].value) if self.is_int(ws['E61'].value) else 0,
                        'cantidad_uzg_cpr': int(ws['E62'].value) if self.is_int(ws['E62'].value) else 0,
                        'cantidad_uzg_ca': int(ws['E63'].value) if self.is_int(ws['E63'].value) else 0,

                        'cantidad_mvih_s_o_n': cantidad_mvih_s_o_n,
                        'cantidad_mhbsag_s_o_n': cantidad_mhbsag_s_o_n,
                        'cantidad_mhepc_s_o_n': cantidad_mhepc_s_o_n,
                        'cantidad_mantihbc_s_o_n': cantidad_mantihbc_s_o_n,
                        'cantidad_mhtlv_s_o_n': cantidad_mhtlv_s_o_n,
                        'cantidad_msifilis_s_o_n': cantidad_msifilis_s_o_n,
                        'cantidad_mchagas_s_o_n': cantidad_mchagas_s_o_n,
                        'cantidad_motros_s_o_n': cantidad_motros_s_o_n,

                        'cantidad_total_donante': total_donantes_line,
                        'cantidad_total_postulante_diferido_excluido': total_postulante_line,
                        'cantidad_total_donacion_segun_grupo_edad': total_donacion_grupo_edad_line,
                        'cantidad_total_donacion_sangre_total': total_sangre_total_line,
                        'cantidad_total_donacion_aferesis': total_donacion_aferesis_line,
                        'cantidad_total_produccion_unidad_sangre': total_produccion_unidad_sangre_line,
                        'cantidad_total_hemocomponente': total_produccion_hemocomponente_line,
                    })

                causa_reaccion_adversa_line = self.env['hemored.ficha_causa_reaccion_adversa.line'].search([('ficha_id', '=', ficha_id.id)])
                total_causa_reaccion_adversa_line = 0
                for line in causa_reaccion_adversa_line:
                    line.write({
                        'cantidad': int(ws[constants.CAUSA_REACCiON_ADVERSA[line.tipo_reaccion_adversa]].value) if self.is_int(ws[constants.CAUSA_REACCiON_ADVERSA[line.tipo_reaccion_adversa]].value) else 0,
                    })
                    total_causa_reaccion_adversa_line += int(ws[constants.CAUSA_REACCiON_ADVERSA[line.tipo_reaccion_adversa]].value) if self.is_int(ws[constants.CAUSA_REACCiON_ADVERSA[line.tipo_reaccion_adversa]].value) else 0

                ficha_id.write({

                    'cantidad_gh1_pact': int(ws['C82'].value) if self.is_int(ws['C82'].value) else 0,
                    'cantidad_gh1_hst': int(ws['D82'].value) if self.is_int(ws['D82'].value) else 0,
                    'cantidad_gh1_hgr': int(ws['E82'].value) if self.is_int(ws['E82'].value) else 0,
                    'cantidad_gh1_hpfc': int(ws['F82'].value) if self.is_int(ws['F82'].value) else 0,
                    'cantidad_gh1_hc': int(ws['G82'].value) if self.is_int(ws['G82'].value) else 0,
                    'cantidad_gh1_hp': int(ws['H82'].value) if self.is_int(ws['H82'].value) else 0,
                    'cantidad_gh1_hap': int(ws['I82'].value) if self.is_int(ws['I82'].value) else 0,
                    'cantidad_gh1_hagr': int(ws['J82'].value) if self.is_int(ws['J82'].value) else 0,
                    'cantidad_gh1_haplasma': int(ws['K82'].value) if self.is_int(ws['K82'].value) else 0,

                    'cantidad_gh2_pact': int(ws['C83'].value) if self.is_int(ws['C83'].value) else 0,
                    'cantidad_gh2_hst': int(ws['D83'].value) if self.is_int(ws['D83'].value) else 0,
                    'cantidad_gh2_hgr': int(ws['E83'].value) if self.is_int(ws['E83'].value) else 0,
                    'cantidad_gh2_hpfc': int(ws['F83'].value) if self.is_int(ws['F83'].value) else 0,
                    'cantidad_gh2_hc': int(ws['G83'].value) if self.is_int(ws['G83'].value) else 0,
                    'cantidad_gh2_hp': int(ws['H83'].value) if self.is_int(ws['H83'].value) else 0,
                    'cantidad_gh2_hap': int(ws['I83'].value) if self.is_int(ws['I83'].value) else 0,
                    'cantidad_gh2_hagr': int(ws['J83'].value) if self.is_int(ws['J83'].value) else 0,
                    'cantidad_gh2_haplasma': int(ws['K83'].value) if self.is_int(ws['K83'].value) else 0,

                    'cantidad_gh3_pact': int(ws['C84'].value) if self.is_int(ws['C84'].value) else 0,
                    'cantidad_gh3_hst': int(ws['D84'].value) if self.is_int(ws['D84'].value) else 0,
                    'cantidad_gh3_hgr': int(ws['E84'].value) if self.is_int(ws['E84'].value) else 0,
                    'cantidad_gh3_hpfc': int(ws['F84'].value) if self.is_int(ws['F84'].value) else 0,
                    'cantidad_gh3_hc': int(ws['G84'].value) if self.is_int(ws['G84'].value) else 0,
                    'cantidad_gh3_hp': int(ws['H84'].value) if self.is_int(ws['H84'].value) else 0,
                    'cantidad_gh3_hap': int(ws['I84'].value) if self.is_int(ws['I84'].value) else 0,
                    'cantidad_gh3_hagr': int(ws['J84'].value) if self.is_int(ws['J84'].value) else 0,
                    'cantidad_gh3_haplasma': int(ws['K84'].value) if self.is_int(ws['K84'].value) else 0,

                    'cantidad_gh4_pact': int(ws['C85'].value) if self.is_int(ws['C85'].value) else 0,
                    'cantidad_gh4_hst': int(ws['D85'].value) if self.is_int(ws['D85'].value) else 0,
                    'cantidad_gh4_hgr': int(ws['E85'].value) if self.is_int(ws['E85'].value) else 0,
                    'cantidad_gh4_hpfc': int(ws['F85'].value) if self.is_int(ws['F85'].value) else 0,
                    'cantidad_gh4_hc': int(ws['G85'].value) if self.is_int(ws['G85'].value) else 0,
                    'cantidad_gh4_hp': int(ws['H85'].value) if self.is_int(ws['H85'].value) else 0,
                    'cantidad_gh4_hap': int(ws['I85'].value) if self.is_int(ws['I85'].value) else 0,
                    'cantidad_gh4_hagr': int(ws['J85'].value) if self.is_int(ws['J85'].value) else 0,
                    'cantidad_gh4_haplasma': int(ws['K85'].value) if self.is_int(ws['K85'].value) else 0,

                    'cantidad_gh5_pact': int(ws['C86'].value) if self.is_int(ws['C86'].value) else 0,
                    'cantidad_gh5_hst': int(ws['D86'].value) if self.is_int(ws['D86'].value) else 0,
                    'cantidad_gh5_hgr': int(ws['E86'].value) if self.is_int(ws['E86'].value) else 0,
                    'cantidad_gh5_hpfc': int(ws['F86'].value) if self.is_int(ws['F86'].value) else 0,
                    'cantidad_gh5_hc': int(ws['G86'].value) if self.is_int(ws['G86'].value) else 0,
                    'cantidad_gh5_hp': int(ws['H86'].value) if self.is_int(ws['H86'].value) else 0,
                    'cantidad_gh5_hap': int(ws['I86'].value) if self.is_int(ws['I86'].value) else 0,
                    'cantidad_gh5_hagr': int(ws['J86'].value) if self.is_int(ws['J86'].value) else 0,
                    'cantidad_gh5_haplasma': int(ws['K86'].value) if self.is_int(ws['K86'].value) else 0,

                    'cantidad_gh6_pact': int(ws['C87'].value) if self.is_int(ws['C87'].value) else 0,
                    'cantidad_gh6_hst': int(ws['D87'].value) if self.is_int(ws['D87'].value) else 0,
                    'cantidad_gh6_hgr': int(ws['E87'].value) if self.is_int(ws['E87'].value) else 0,
                    'cantidad_gh6_hpfc': int(ws['F87'].value) if self.is_int(ws['F87'].value) else 0,
                    'cantidad_gh6_hc': int(ws['G87'].value) if self.is_int(ws['G87'].value) else 0,
                    'cantidad_gh6_hp': int(ws['H87'].value) if self.is_int(ws['H87'].value) else 0,
                    'cantidad_gh6_hap': int(ws['I87'].value) if self.is_int(ws['I87'].value) else 0,
                    'cantidad_gh6_hagr': int(ws['J87'].value) if self.is_int(ws['J87'].value) else 0,
                    'cantidad_gh6_haplasma': int(ws['K87'].value) if self.is_int(ws['K87'].value) else 0,

                    'cantidad_gh7_pact': int(ws['C88'].value) if self.is_int(ws['C88'].value) else 0,
                    'cantidad_gh7_hst': int(ws['D88'].value) if self.is_int(ws['D88'].value) else 0,
                    'cantidad_gh7_hgr': int(ws['E88'].value) if self.is_int(ws['E88'].value) else 0,
                    'cantidad_gh7_hpfc': int(ws['F88'].value) if self.is_int(ws['F88'].value) else 0,
                    'cantidad_gh7_hc': int(ws['G88'].value) if self.is_int(ws['G88'].value) else 0,
                    'cantidad_gh7_hp': int(ws['H88'].value) if self.is_int(ws['H88'].value) else 0,
                    'cantidad_gh7_hap': int(ws['I88'].value) if self.is_int(ws['I88'].value) else 0,
                    'cantidad_gh7_hagr': int(ws['J88'].value) if self.is_int(ws['J88'].value) else 0,
                    'cantidad_gh7_haplasma': int(ws['K88'].value) if self.is_int(ws['K88'].value) else 0,

                    'cantidad_ds_hgr': int(ws['D94'].value) if self.is_int(ws['D94'].value) else 0,
                    'cantidad_ds_hpfc': int(ws['E94'].value) if self.is_int(ws['E94'].value) else 0,
                    'cantidad_ds_hp': int(ws['F94'].value) if self.is_int(ws['F94'].value) else 0,
                    'cantidad_ds_hap': int(ws['G94'].value) if self.is_int(ws['G94'].value) else 0,

                    'cantidad_da_hgr': int(ws['D95'].value) if self.is_int(ws['D95'].value) else 0,
                    'cantidad_da_hpfc': int(ws['E95'].value) if self.is_int(ws['E95'].value) else 0,
                    'cantidad_da_hp': int(ws['F95'].value) if self.is_int(ws['F95'].value) else 0,
                    'cantidad_da_hap': int(ws['G95'].value) if self.is_int(ws['G95'].value) else 0,

                    'cantidad_tur_hst': int(ws['D100'].value) if self.is_int(ws['D100'].value) else 0,
                    'cantidad_tur_hgr': int(ws['E100'].value) if self.is_int(ws['E100'].value) else 0,
                    'cantidad_tur_hpfc': int(ws['F100'].value) if self.is_int(ws['F100'].value) else 0,
                    'cantidad_tur_hc': int(ws['G100'].value) if self.is_int(ws['G100'].value) else 0,
                    'cantidad_tur_hp': int(ws['H100'].value) if self.is_int(ws['H100'].value) else 0,
                    'cantidad_tur_hap': int(ws['I100'].value) if self.is_int(ws['I100'].value) else 0,
                    'cantidad_tur_hagr': int(ws['J100'].value) if self.is_int(ws['J100'].value) else 0,

                    'cantidad_tut_hst': int(ws['D101'].value) if self.is_int(ws['D101'].value) else 0,
                    'cantidad_tut_hgr': int(ws['E101'].value) if self.is_int(ws['E101'].value) else 0,
                    'cantidad_tut_hpfc': int(ws['F101'].value) if self.is_int(ws['F101'].value) else 0,
                    'cantidad_tut_hc': int(ws['G101'].value) if self.is_int(ws['G101'].value) else 0,
                    'cantidad_tut_hp': int(ws['H101'].value) if self.is_int(ws['H101'].value) else 0,
                    'cantidad_tut_hap': int(ws['I101'].value) if self.is_int(ws['I101'].value) else 0,
                    'cantidad_tut_hagr': int(ws['J101'].value) if self.is_int(ws['J101'].value) else 0,

                    'cantidad_cv_hst': int(ws['C129'].value) if self.is_int(ws['C129'].value) else 0,
                    'cantidad_cv_hgr': int(ws['D129'].value) if self.is_int(ws['D129'].value) else 0,
                    'cantidad_cv_hpfc': int(ws['E129'].value) if self.is_int(ws['E129'].value) else 0,
                    'cantidad_cv_hc': int(ws['F129'].value) if self.is_int(ws['F129'].value) else 0,
                    'cantidad_cv_hp': int(ws['G129'].value) if self.is_int(ws['G129'].value) else 0,
                    'cantidad_cv_hap': int(ws['H129'].value) if self.is_int(ws['H129'].value) else 0,
                    'cantidad_cv_hagr': int(ws['I129'].value) if self.is_int(ws['I129'].value) else 0,
                    'cantidad_cv_haplasma': int(ws['J129'].value) if self.is_int(ws['J129'].value) else 0,

                    'cantidad_cmitt_hst': int(ws['C130'].value) if self.is_int(ws['C130'].value) else 0,
                    'cantidad_cmitt_hgr': int(ws['D130'].value) if self.is_int(ws['D130'].value) else 0,
                    'cantidad_cmitt_hpfc': int(ws['E130'].value) if self.is_int(ws['E130'].value) else 0,
                    'cantidad_cmitt_hc': int(ws['F130'].value) if self.is_int(ws['F130'].value) else 0,
                    'cantidad_cmitt_hp': int(ws['G130'].value) if self.is_int(ws['G130'].value) else 0,
                    'cantidad_cmitt_hap': int(ws['H130'].value) if self.is_int(ws['H130'].value) else 0,
                    'cantidad_cmitt_hagr': int(ws['I130'].value) if self.is_int(ws['I130'].value) else 0,
                    'cantidad_cmitt_haplasma': int(ws['J130'].value) if self.is_int(ws['J130'].value) else 0,

                    'cantidad_ca_hst': int(ws['C131'].value) if self.is_int(ws['C131'].value) else 0,
                    'cantidad_ca_hgr': int(ws['D131'].value) if self.is_int(ws['D131'].value) else 0,
                    'cantidad_ca_hpfc': int(ws['E131'].value) if self.is_int(ws['E131'].value) else 0,
                    'cantidad_ca_hc': int(ws['F131'].value) if self.is_int(ws['F131'].value) else 0,
                    'cantidad_ca_hp': int(ws['G131'].value) if self.is_int(ws['G131'].value) else 0,
                    'cantidad_ca_hap': int(ws['H131'].value) if self.is_int(ws['H131'].value) else 0,
                    'cantidad_ca_hagr': int(ws['I131'].value) if self.is_int(ws['I131'].value) else 0,
                    'cantidad_ca_haplasma': int(ws['J131'].value) if self.is_int(ws['J131'].value) else 0,

                    'cantidad_ct_hst': int(ws['C132'].value) if self.is_int(ws['C132'].value) else 0,
                    'cantidad_ct_hgr': int(ws['D132'].value) if self.is_int(ws['D132'].value) else 0,
                    'cantidad_ct_hpfc': int(ws['E132'].value) if self.is_int(ws['E132'].value) else 0,
                    'cantidad_ct_hc': int(ws['F132'].value) if self.is_int(ws['F132'].value) else 0,
                    'cantidad_ct_hp': int(ws['G132'].value) if self.is_int(ws['G132'].value) else 0,
                    'cantidad_ct_hap': int(ws['H132'].value) if self.is_int(ws['H132'].value) else 0,
                    'cantidad_ct_hagr': int(ws['I132'].value) if self.is_int(ws['I132'].value) else 0,
                    'cantidad_ct_haplasma': int(ws['J132'].value) if self.is_int(ws['J132'].value) else 0,

                    'cantidad_cp_hst': int(ws['C133'].value) if self.is_int(ws['C133'].value) else 0,
                    'cantidad_cp_hgr': int(ws['D133'].value) if self.is_int(ws['D133'].value) else 0,
                    'cantidad_cp_hpfc': int(ws['E133'].value) if self.is_int(ws['E133'].value) else 0,
                    'cantidad_cp_hc': int(ws['F133'].value) if self.is_int(ws['F133'].value) else 0,
                    'cantidad_cp_hp': int(ws['G133'].value) if self.is_int(ws['G133'].value) else 0,
                    'cantidad_cp_hap': int(ws['H133'].value) if self.is_int(ws['H133'].value) else 0,
                    'cantidad_cp_hagr': int(ws['I133'].value) if self.is_int(ws['I133'].value) else 0,
                    'cantidad_cp_haplasma': int(ws['J133'].value) if self.is_int(ws['J133'].value) else 0,

                    'cantidad_co_hst': int(ws['C134'].value) if self.is_int(ws['C134'].value) else 0,
                    'cantidad_co_hgr': int(ws['D134'].value) if self.is_int(ws['D134'].value) else 0,
                    'cantidad_co_hpfc': int(ws['E134'].value) if self.is_int(ws['E134'].value) else 0,
                    'cantidad_co_hc': int(ws['F134'].value) if self.is_int(ws['F134'].value) else 0,
                    'cantidad_co_hp': int(ws['G134'].value) if self.is_int(ws['G134'].value) else 0,
                    'cantidad_co_hap': int(ws['H134'].value) if self.is_int(ws['H134'].value) else 0,
                    'cantidad_co_hagr': int(ws['I134'].value) if self.is_int(ws['I134'].value) else 0,
                    'cantidad_co_haplasma': int(ws['J134'].value) if self.is_int(ws['J134'].value) else 0,

                    'cantidad_total_causa_reaccion_adversa': total_causa_reaccion_adversa_line,
                })

            elif not fields.Date.today() <= ficha_id.fin_fecha_registro:
                raise ValidationError('La Ficha ha vencido, el Fin de Fecha Registro es menor a la fecha actual, si requiere más tiempo, comuníquese a la central')
            else:
                raise ValidationError('Error al Cargar el Excel, los datos de Validación no coinciden')
            os.unlink(fname)
        else:
            raise ValidationError('Error Ficha se Encuentra en estado Validado')

    @api.one
    @api.constrains('ficha_estadistica')
    def _check_ficha_estadistica(self):
        if not (self.txt_ficha_estadistica.endswith('.xls') or self.txt_ficha_estadistica.endswith('.xlsx')):
            raise ValidationError('Tipo de Archivo NO Válido')
