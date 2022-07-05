# -*- coding: utf-8 -*-
import base64
import os
import tempfile
from odoo import fields, api, models
from odoo.exceptions import ValidationError
from openpyxl import load_workbook

from ..models import constants


class UpdateExcel(models.TransientModel):
    _name = 'hemored_update_supervision.updateexcel'

    ficha_supervision = fields.Binary(
        string='Ficha de Supervisión',
        attachment=True,
    )
    txt_ficha_supervision = fields.Char(
        string='Nombre Ficha de Supervisión'
    )

    @api.multi
    def is_int(self, valor):
        try:
            if valor is None:
                return False
            else:
                int(valor)
                return True
        except ValueError:
            return False

    @api.multi
    def cargar_excel_ficha(self):
        ficha = self.env.context.get('active_id')
        ficha_id = self.env['hemored_supervision.record_supervision'].browse(ficha)
        if ficha_id.state != constants.TERMINADO:
            data = base64.decodestring(self.ficha_supervision)
            fobj = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            fname = fobj.name
            fobj.write(data)
            fobj.close()
            book = load_workbook(fname)
            sheets = book.sheetnames
            ws = book[sheets[0]]
            if ficha_id.banco_id.renipress_id.codigo_eess == ws['E9'].value:
                rg_6 = False
                p_rg_6 = False
                g_rg_6 = False
                rg_14 = False
                p_rg_14 = False
                g_rg_14 = False
                rg_17 = False
                p_rg_17 = False
                g_rg_17 = False
                t_rg_22 = False
                rg_26 = False
                p_rg_26 = False
                g_rg_26 = False
                pros_2 = False
                p_pros_2 = False
                g_pros_2 = False
                pros_3 = False
                p_pros_3 = False
                g_pros_3 = False
                t_pros_4 = False
                pros_5 = False
                p_pros_5 = False
                g_pros_5 = False
                pros_6 = False
                p_pros_6 = False
                g_pros_6 = False
                t_pros_7 = False
                p_pros_7 = False
                g_pros_7 = False
                t_pros_8 = False
                p_pros_8 = False
                g_pros_8 = False
                pros_9 = False
                p_pros_9 = False
                g_pros_9 = False
                pros_10 = False
                p_pros_10 = False
                g_pros_10 = False
                pros_11 = False
                p_pros_11 = False
                g_pros_11 = False
                t_pros_12 = False
                p_pros_12 = False
                g_pros_12 = False
                pros_13 = False
                p_pros_13 = False
                g_pros_13 = False
                pros_14 = False
                p_pros_14 = False
                g_pros_14 = False
                pros_15 = False
                p_pros_15 = False
                g_pros_15 = False
                pros_16 = False
                p_pros_16 = False
                g_pros_16 = False
                pros_17 = False
                p_pros_17 = False
                g_pros_17 = False
                t_pros_17 = False
                pros_18 = False
                p_pros_18 = False
                g_pros_18 = False
                t_pros_19 = False
                pros_21 = False
                p_pros_21 = False
                g_pros_21 = False
                pros_22 = False
                p_pros_22 = False
                g_pros_22 = False
                t_pros_23 = False
                t_pros_24 = False
                pros_33 = False
                pros_25 = False
                pros_26 = False
                pros_27 = False
                pros_28 = False
                pros_29 = False
                pros_30 = False
                pros_31 = False
                pros_32 = False
                pros_35 = False
                p_pros_35 = False
                g_pros_35 = False
                pros_36 = False
                p_pros_36 = False
                g_pros_36 = False
                pros_37 = False
                p_pros_37 = False
                g_pros_37 = False
                t_pros_38 = False
                pros_38 = False
                p_pros_38 = False
                g_pros_38 = False
                t_pros_39 = False
                pros_39 = False
                p_pros_39 = False
                g_pros_39 = False
                t_pros_40 = False
                pros_40 = False
                p_pros_40 = False
                g_pros_40 = False
                t_pros_41 = False
                pros_41 = False
                p_pros_41 = False
                g_pros_41 = False
                t_pros_42 = False
                pros_42 = False
                p_pros_42 = False
                g_pros_42 = False
                t_pros_43 = False
                pros_43 = False
                p_pros_43 = False
                g_pros_43 = False
                t_pros_44 = False
                pros_44 = False
                p_pros_44 = False
                g_pros_44 = False
                pros_45 = False
                t_pros_47 = False
                pros_49 = False
                p_pros_49 = False
                g_pros_49 = False
                pros_50 = False
                p_pros_50 = False
                g_pros_50 = False
                t_pros_50 = False
                pros_51 = False
                p_pros_51 = False
                g_pros_51 = False
                pros_52 = False
                p_pros_52 = False
                g_pros_52 = False
                pros_53 = False
                p_pros_53 = False
                g_pros_53 = False
                r_pros_53 = False
                p_r_pros_53 = False
                g_r_pros_53 = False
                pros_54 = False
                p_pros_54 = False
                g_pros_54 = False
                pros_55 = False
                p_pros_55 = False
                g_pros_55 = False
                t_pros_56 = False
                pros_57 = False
                p_pros_57 = False
                g_pros_57 = False
                pros_58 = False
                p_pros_58 = False
                g_pros_58 = False
                pros_59 = False
                p_pros_59 = False
                g_pros_59 = False
                pros_60 = False
                p_pros_60 = False
                g_pros_60 = False
                pros_61 = False
                p_pros_61 = False
                g_pros_61 = False
                pros_62 = False
                p_pros_62 = False
                g_pros_62 = False
                pros_63 = False
                p_pros_63 = False
                g_pros_63 = False
                pros_64 = False
                p_pros_64 = False
                g_pros_64 = False
                pros_65 = False
                p_pros_65 = False
                g_pros_65 = False
                pros_66 = False
                p_pros_66 = False
                g_pros_66 = False
                pros_68 = False
                t_pros_68 = False
                pros_81 = False
                pros_82 = False
                pros_83 = False
                pros_84 = False
                pros_85 = False
                p_pros_85 = False
                g_pros_85 = False
                pros_87 = False
                pros_88 = False
                pros_89 = False
                pros_90 = False
                pros_91 = False
                pros_92 = False
                pros_93 = False
                p_pros_93 = False
                g_pros_93 = False
                t_pros_93 = False
                pros_99 = False
                p_pros_99 = False
                g_pros_99 = False
                pros_100 = False
                p_pros_100 = False
                g_pros_100 = False
                pros_101 = False
                p_pros_101 = False
                g_pros_101 = False
                pros_102 = False
                p_pros_102 = False
                g_pros_102 = False
                pros_103 = False
                p_pros_103 = False
                g_pros_103 = False
                cali_4 = False
                cali_5 = False
                cali_6 = False
                cali_7 = False
                cali_8 = False
                clasf_bio_3 = False
                p_clasf_bio_3 = False
                g_clasf_bio_3 = False
                t_clasf_bio_5 = False
                clasf_bio_6 = False
                p_clasf_bio_6 = False
                g_clasf_bio_6 = False
                clasf_bio_7 = False
                p_clasf_bio_7 = False
                g_clasf_bio_7 = False
                clasf_bio_18 = False
                p_clasf_bio_18 = False
                g_clasf_bio_18 = False
                clasf_bio_19 = False
                p_clasf_bio_19 = False
                g_clasf_bio_19 = False
                equip_8 = False
                p_equip_8 = False
                g_equip_8 = False
                rel_c_equip_1 = False
                rel_p_equip_1 = False
                rel_c_equip_2 = False
                rel_p_equip_2 = False
                rel_c_equip_6 = False
                rel_p_equip_6 = False
                rel_c_equip_9 = False
                rel_p_equip_9 = False
                rel_c_equip_10 = False
                rel_p_equip_10 = False
                rel_c_equip_11 = False
                rel_p_equip_11 = False
                rel_c_equip_13 = False
                rel_p_equip_13 = False
                rel_c_equip_24 = False
                rel_p_equip_24 = False
                rel_c_equip_25 = False
                rel_p_equip_25 = False
                rel_c_equip_26 = False
                rel_p_equip_26 = False
                rel_c_equip_33 = False
                rel_p_equip_33 = False
                rel_c_equip_34 = False
                rel_p_equip_34 = False
                infra_instal_5 = False
                p_infra_instal_5 = False
                g_infra_instal_5 = False
                infra_instal_7 = False
                p_infra_instal_7 = False
                g_infra_instal_7 = False
                infra_instal_8 = False
                p_infra_instal_8 = False
                g_infra_instal_8 = False
                infra_instal_9 = False
                p_infra_instal_9 = False
                g_infra_instal_9 = False
                infra_instal_10 = False
                infra_instal_12 = False
                p_infra_instal_12 = False
                g_infra_instal_12 = False
                infra_instal_13 = False
                p_infra_instal_13 = False
                g_infra_instal_13 = False
                infra_instal_15 = False
                p_infra_instal_15 = False
                g_infra_instal_15 = False
                infra_instal_16 = False
                p_infra_instal_16 = False
                g_infra_instal_16 = False
                rg_1 = False
                p_rg_1 = False
                g_rg_1 = False
                rg_2 = False
                p_rg_2 = False
                g_rg_2 = False
                rg_4 = False
                p_rg_4 = False
                g_rg_4 = False
                rg_5 = False
                p_rg_5 = False
                g_rg_5 = False
                rg_8 = False
                p_rg_8 = False
                g_rg_8 = False
                rg_9 = False
                p_rg_9 = False
                g_rg_9 = False
                rg_10 = False
                p_rg_10 = False
                g_rg_10 = False
                rg_11 = False
                p_rg_11 = False
                g_rg_11 = False
                rg_12 = False
                p_rg_12 = False
                g_rg_12 = False
                rg_13 = False
                p_rg_13 = False
                g_rg_13 = False
                rg_15 = False
                p_rg_15 = False
                g_rg_15 = False
                rg_16 = False
                p_rg_16 = False
                g_rg_16 = False
                rg_18 = False
                p_rg_18 = False
                g_rg_18 = False
                rg_19 = False
                p_rg_19 = False
                g_rg_19 = False
                rg_21 = False
                p_rg_21 = False
                g_rg_21 = False
                rg_23 = False
                p_rg_23 = False
                g_rg_23 = False
                rg_24 = False
                p_rg_24 = False
                g_rg_24 = False
                rg_25 = False
                p_rg_25 = False
                g_rg_25 = False
                rg_28 = False
                p_rg_28 = False
                g_rg_28 = False
                rg_29 = False
                p_rg_29 = False
                g_rg_29 = False
                rg_30 = False
                p_rg_30 = False
                g_rg_30 = False
                rh_1 = False
                p_rh_1 = False
                g_rh_1 = False
                rh_2 = False
                p_rh_2 = False
                g_rh_2 = False
                rh_3 = False
                p_rh_3 = False
                g_rh_3 = False
                rh_4 = False
                p_rh_4 = False
                g_rh_4 = False
                rh_5 = False
                p_rh_5 = False
                g_rh_5 = False
                rh_6 = False
                p_rh_6 = False
                g_rh_6 = False
                rh_7 = False
                p_rh_7 = False
                g_rh_7 = False
                rh_8 = False
                p_rh_8 = False
                g_rh_8 = False
                rh_9 = False
                p_rh_9 = False
                g_rh_9 = False
                rh_10 = False
                p_rh_10 = False
                g_rh_10 = False
                pros_34 = False
                p_pros_34 = False
                g_pros_34 = False
                pros_69 = False
                p_pros_69 = False
                g_pros_69 = False
                pros_70 = False
                p_pros_70 = False
                g_pros_70 = False
                pros_73 = False
                p_pros_73 = False
                g_pros_73 = False
                pros_74 = False
                p_pros_74 = False
                g_pros_74 = False
                pros_75 = False
                p_pros_75 = False
                g_pros_75 = False
                pros_76 = False
                p_pros_76 = False
                g_pros_76 = False
                pros_77 = False
                p_pros_77 = False
                g_pros_77 = False
                pros_78 = False
                p_pros_78 = False
                g_pros_78 = False
                pros_79 = False
                p_pros_79 = False
                g_pros_79 = False
                pros_80 = False
                p_pros_80 = False
                g_pros_80 = False
                pros_94 = False
                p_pros_94 = False
                g_pros_94 = False
                pros_95 = False
                p_pros_95 = False
                g_pros_95 = False
                pros_96 = False
                pros_97 = False
                pros_98 = False
                p_pros_98 = False
                g_pros_98 = False
                cali_1 = False
                cali_2 = False
                cali_3 = False
                alm_6 = False
                p_alm_6 = False
                g_alm_6 = False
                alm_7 = False
                p_alm_7 = False
                g_alm_7 = False
                alm_8 = False
                p_alm_8 = False
                g_alm_8 = False
                sis_inform_1 = False
                sis_inform_5 = False
                clasf_bio_2 = False
                p_clasf_bio_2 = False
                g_clasf_bio_2 = False
                clasf_bio_8 = False
                p_clasf_bio_8 = False
                g_clasf_bio_8 = False
                clasf_bio_14 = False
                p_clasf_bio_14 = False
                g_clasf_bio_14 = False
                clasf_bio_15 = False
                p_clasf_bio_15 = False
                g_clasf_bio_15 = False
                equip_1 = False
                p_equip_1 = False
                g_equip_1 = False
                equip_2 = False
                p_equip_2 = False
                g_equip_2 = False
                equip_3 = False
                p_equip_3 = False
                g_equip_3 = False
                equip_4 = False
                p_equip_4 = False
                g_equip_4 = False
                equip_5 = False
                p_equip_5 = False
                g_equip_5 = False
                equip_6 = False
                p_equip_6 = False
                g_equip_6 = False
                equip_7 = False
                p_equip_7 = False
                g_equip_7 = False
                equip_9 = False
                p_equip_9 = False
                g_equip_9 = False
                equip_10 = False
                p_equip_10 = False
                g_equip_10 = False
                equip_11 = False
                p_equip_11 = False
                g_equip_11 = False
                rel_c_equip_3 = False
                rel_p_equip_3 = False
                rel_c_equip_4 = False
                rel_p_equip_4 = False
                rel_c_equip_5 = False
                rel_p_equip_5 = False
                rel_c_equip_7 = False
                rel_p_equip_7 = False
                rel_c_equip_8 = False
                rel_p_equip_8 = False
                rel_c_equip_12 = False
                rel_p_equip_12 = False
                rel_c_equip_14 = False
                rel_p_equip_14 = False
                rel_c_equip_15 = False
                rel_p_equip_15 = False
                rel_c_equip_16 = False
                rel_p_equip_16 = False
                rel_c_equip_17 = False
                rel_p_equip_17 = False
                rel_c_equip_18 = False
                rel_p_equip_18 = False
                rel_c_equip_19 = False
                rel_p_equip_19 = False
                rel_c_equip_20 = False
                rel_p_equip_20 = False
                rel_c_equip_21 = False
                rel_p_equip_21 = False
                rel_c_equip_22 = False
                rel_p_equip_22 = False
                rel_c_equip_23 = False
                rel_p_equip_23 = False
                rel_c_equip_31 = False
                rel_p_equip_31 = False
                rel_c_equip_32 = False
                rel_p_equip_32 = False
                rel_c_equip_27 = False
                rel_p_equip_27 = False
                rel_c_equip_28 = False
                rel_p_equip_28 = False
                rel_c_equip_29 = False
                rel_p_equip_29 = False
                rel_c_equip_30 = False
                rel_p_equip_30 = False
                infra_instal_6 = False
                p_infra_instal_6 = False
                g_infra_instal_6 = False
                infra_instal_11 = False
                p_infra_instal_11 = False
                g_infra_instal_11 = False
                infra_instal_14 = False
                p_infra_instal_14 = False
                g_infra_instal_14 = False
                infra_instal_17 = False
                p_infra_instal_17 = False
                g_infra_instal_17 = False
                infra_instal_18 = False
                p_infra_instal_18 = False
                g_infra_instal_18 = False
                infra_instal_19 = False
                p_infra_instal_19 = False
                g_infra_instal_19 = False
                infra_instal_20 = False
                p_infra_instal_20 = False
                g_infra_instal_20 = False
                infra_instal_21 = False
                p_infra_instal_21 = False
                g_infra_instal_21 = False
                infra_instal_22 = False
                p_infra_instal_22 = False
                g_infra_instal_22 = False
                infra_instal_23 = False
                p_infra_instal_23 = False
                g_infra_instal_23 = False
                infra_instal_24 = False
                infra_instal_25 = False
                infra_instal_26 = False
                p_infra_instal_26 = False
                g_infra_instal_26 = False
                infra_instal_27 = False
                infra_instal_28 = False
                bio_s_2 = False
                p_bio_s_2 = False
                g_bio_s_2 = False
                bio_s_3 = False
                bio_s_4 = False
                bio_s_5 = False
                p_bio_s_5 = False
                g_bio_s_5 = False
                if ficha_id.banco_id.tipo_banco == '2':
                    if ws['H53'].value == 'SI':
                        rg_6 = constants.SI
                        p_rg_6 = 4
                        g_rg_6 = constants.NULO
                    elif ws['H53'].value == 'NO':
                        rg_6 = constants.NO
                        p_rg_6 = 0
                        g_rg_6 = constants.LEVE
                    if ws['H60'].value == 'SI':
                        rg_14 = constants.SI
                        p_rg_14 = 4
                        g_rg_14 = constants.NULO
                    elif ws['H60'].value == 'NO':
                        rg_14 = constants.NO
                        p_rg_14 = 0
                        g_rg_14 = constants.MODERADO
                    if ws['H63'].value == 'SI':
                        rg_17 = constants.SI
                        p_rg_17 = 4
                        g_rg_17 = constants.NULO
                    elif ws['H63'].value == 'NO':
                        rg_17 = constants.NO
                        p_rg_17 = 0
                        g_rg_17 = constants.LEVE
                    if ws['H72'].value == 'SI':
                        t_rg_22 = constants.UNO
                    elif ws['H73'].value == 'SI':
                        t_rg_22 = constants.DOS
                    elif ws['H74'].value == 'SI':
                        t_rg_22 = constants.TRES
                    elif ws['H75'].value == 'SI':
                        t_rg_22 = constants.CUATRO
                    elif ws['H76'].value == 'SI':
                        t_rg_22 = constants.CINCO
                    elif ws['H77'].value == 'SI':
                        t_rg_22 = constants.SEIS
                    else:
                        t_rg_22 = False
                    if ws['H81'].value == 'SI':
                        rg_26 = constants.SI
                        p_rg_26 = 4
                        g_rg_26 = constants.NULO
                    elif ws['H81'].value == 'NO':
                        rg_26 = constants.NO
                        p_rg_26 = 0
                        g_rg_26 = constants.MODERADO
                    if ws['H103'].value == 'SI':
                        pros_2 = constants.SI
                        p_pros_2 = 5
                        g_pros_2 = constants.NULO
                    elif ws['H103'].value == 'NO':
                        pros_2 = constants.NO
                        p_pros_2 = 0
                        g_pros_2 = constants.MODERADO
                    if ws['H104'].value == 'SI':
                        pros_3 = constants.SI
                        p_pros_3 = 5
                        g_pros_3 = constants.NULO
                    elif ws['H104'].value == 'NO':
                        pros_3 = constants.NO
                        p_pros_3 = 0
                        g_pros_3 = constants.MODERADO
                    if ws['H105'].value == 'SI':
                        t_pros_4 = constants.DNI
                    elif ws['H106'].value == 'SI':
                        t_pros_4 = constants.PASAPORTE
                    elif ws['H107'].value == 'SI':
                        t_pros_4 = constants.CARNET_EXTRANJERIA
                    else:
                        t_pros_4 = False
                    if ws['H108'].value == 'SI':
                        pros_5 = constants.SI
                        p_pros_5 = 5
                        g_pros_5 = constants.NULO
                    elif ws['H108'].value == 'NO':
                        pros_5 = constants.NO
                        p_pros_5 = 0
                        g_pros_5 = constants.MODERADO
                    if ws['H109'].value == 'SI':
                        pros_6 = constants.SI
                        p_pros_6 = 5
                        g_pros_6 = constants.NULO
                    elif ws['H109'].value == 'NO':
                        pros_6 = constants.NO
                        p_pros_6 = 0
                        g_pros_6 = constants.MODERADO
                    if ws['H110'].value == 'SI':
                        t_pros_7 = constants.ENTREVISTA
                        p_pros_7 = 4
                        g_pros_7 = constants.NULO
                    elif ws['H111'].value == 'SI':
                        t_pros_7 = constants.AUTOAPLICADO
                        p_pros_7 = 0
                        g_pros_7 = constants.MODERADO
                    else:
                        t_pros_7 = False
                        p_pros_7 = 0
                        g_pros_7 = False
                    if ws['H112'].value == 'SI':
                        t_pros_8 = constants.MEDICO
                        p_pros_8 = 5
                        g_pros_8 = constants.NULO
                    elif ws['H113'].value == 'SI':
                        t_pros_8 = constants.TECNOLOGO
                        p_pros_8 = 4
                        g_pros_8 = constants.NULO
                    elif ws['H114'].value == 'SI':
                        t_pros_8 = constants.OTRO
                        p_pros_8 = 3
                        g_pros_8 = constants.LEVE
                    else:
                        t_pros_8 = False
                        p_pros_8 = 0
                        g_pros_8 = constants.NULO
                    if ws['H115'].value == 'SI':
                        pros_9 = constants.SI
                        p_pros_9 = 5
                        g_pros_9 = constants.NULO
                    elif ws['H115'].value == 'NO':
                        pros_9 = constants.NO
                        p_pros_9 = 0
                        g_pros_9 = constants.GRAVE
                    if ws['H116'].value == 'SI':
                        pros_10 = constants.SI
                        p_pros_10 = 4
                        g_pros_10 = constants.NULO
                    elif ws['H116'].value == 'NO':
                        pros_10 = constants.NO
                        p_pros_10 = 0
                        g_pros_10 = constants.MODERADO
                    if ws['H117'].value == 'SI':
                        pros_11 = constants.SI
                        p_pros_11 = 4
                        g_pros_11 = constants.NULO
                    elif ws['H117'].value == 'NO':
                        pros_11 = constants.NO
                        p_pros_11 = 0
                        g_pros_11 = constants.MODERADO
                    if ws['H118'].value == 'SI':
                        t_pros_12 = constants.COMPLETO
                        p_pros_12 = 5
                        g_pros_12 = constants.NULO
                    elif ws['H119'].value == 'SI':
                        t_pros_12 = constants.INCOMPLETO
                        p_pros_12 = 0
                        g_pros_12 = constants.MODERADO
                    else:
                        t_pros_12 = False
                        p_pros_12 = 0
                        g_pros_12 = False
                    if ws['H120'].value == 'SI':
                        pros_13 = constants.SI
                        p_pros_13 = 4
                        g_pros_13 = constants.NULO
                    elif ws['H120'].value == 'NO':
                        pros_13 = constants.NO
                        p_pros_13 = 0
                        g_pros_13 = constants.MODERADO
                    if ws['H121'].value == 'SI':
                        pros_14 = constants.SI
                        p_pros_14 = 4
                        g_pros_14 = constants.NULO
                    elif ws['H121'].value == 'NO':
                        pros_14 = constants.NO
                        p_pros_14 = 0
                        g_pros_14 = constants.MODERADO
                    if ws['H122'].value == 'SI':
                        pros_15 = constants.SI
                        p_pros_15 = 4
                        g_pros_15 = constants.NULO
                    elif ws['H122'].value == 'NO':
                        pros_15 = constants.NO
                        p_pros_15 = 0
                        g_pros_15 = constants.MODERADO
                    if ws['H123'].value == 'SI':
                        pros_16 = constants.SI
                        p_pros_16 = 5
                        g_pros_16 = constants.NULO
                    elif ws['H123'].value == 'NO':
                        pros_16 = constants.NO
                        p_pros_16 = 0
                        g_pros_16 = constants.GRAVE
                    if ws['H124'].value == 'SI':
                        pros_17 = constants.SI
                        p_pros_17 = 5
                        g_pros_17 = constants.NULO
                    elif ws['H124'].value == 'NO':
                        pros_17 = constants.NO
                        p_pros_17 = 0
                        g_pros_17 = constants.GRAVE
                    if ws['H125'].value == 'SI':
                        t_pros_17 = constants.PRE_DONACION
                    elif ws['H126'].value == 'SI':
                        t_pros_17 = constants.POST_DONACION
                    elif ws['H127'].value == 'SI':
                        t_pros_17 = constants.AMBOS
                    else:
                        t_pros_17 = False
                    if ws['H128'].value == 'SI':
                        pros_18 = constants.SI
                        p_pros_18 = 4
                        g_pros_18 = constants.NULO
                    elif ws['H128'].value == 'NO':
                        pros_18 = constants.NO
                        p_pros_18 = 0
                        g_pros_18 = constants.MODERADO
                    if ws['H129'].value == 'SI':
                        t_pros_19 = constants.EMAIL
                    elif ws['H130'].value == 'SI':
                        t_pros_19 = constants.PRESENCIAL
                    else:
                        t_pros_19 = False
                    if ws['H133'].value == 'SI':
                        pros_21 = constants.SI
                        p_pros_21 = 3
                        g_pros_21 = constants.NULO
                    elif ws['H133'].value == 'NO':
                        pros_21 = constants.NO
                        p_pros_21 = 0
                        g_pros_21 = constants.MODERADO
                    if ws['H134'].value == 'SI':
                        pros_22 = constants.SI
                        p_pros_22 = 5
                        g_pros_22 = constants.NULO
                    elif ws['H134'].value == 'NO':
                        pros_22 = constants.NO
                        p_pros_22 = 0
                        g_pros_22 = constants.GRAVE
                    if ws['H135'].value == 'SI':
                        t_pros_23 = constants.MANUAL
                    elif ws['H136'].value == 'SI':
                        t_pros_23 = constants.SELLADOR_ELECTRONICO
                    else:
                        t_pros_23 = False
                    if ws['H137'].value == 'SI':
                        t_pros_24 = constants.DOBLE
                    elif ws['H138'].value == 'SI':
                        t_pros_24 = constants.TRIPLE
                    elif ws['H139'].value == 'SI':
                        t_pros_24 = constants.CUADRUPLE
                    else:
                        t_pros_24 = False
                    if ws['H140'].value == 'SI':
                        pros_33 = constants.SI
                    elif ws['H140'].value == 'NO':
                        pros_33 = constants.NO
                    if ws['H142'].value == 'SI':
                        pros_25 = constants.SI
                    elif ws['H142'].value == 'NO':
                        pros_25 = constants.NO
                    if ws['H143'].value == 'SI':
                        pros_26 = constants.SI
                    elif ws['H143'].value == 'NO':
                        pros_26 = constants.NO
                    if ws['H144'].value == 'SI':
                        pros_27 = constants.SI
                    elif ws['H144'].value == 'NO':
                        pros_27 = constants.NO
                    if ws['H145'].value == 'SI':
                        pros_28 = constants.SI
                    elif ws['H145'].value == 'NO':
                        pros_28 = constants.NO
                    if ws['H146'].value == 'SI':
                        pros_29 = constants.SI
                    elif ws['H146'].value == 'NO':
                        pros_29 = constants.NO
                    if ws['H147'].value == 'SI':
                        pros_30 = constants.SI
                    elif ws['H147'].value == 'NO':
                        pros_30 = constants.NO
                    if ws['H148'].value == 'SI':
                        pros_31 = constants.SI
                    elif ws['H148'].value == 'NO':
                        pros_31 = constants.NO
                    if ws['H149'].value == 'SI':
                        pros_32 = constants.SI
                    elif ws['H149'].value == 'NO':
                        pros_32 = constants.NO
                    if ws['H154'].value == 'SI':
                        pros_35 = constants.SI
                        p_pros_35 = 4
                        g_pros_35 = constants.NULO
                    elif ws['H154'].value == 'NO':
                        pros_35 = constants.NO
                        p_pros_35 = 0
                        g_pros_35 = constants.MODERADO
                    if ws['H155'].value == 'SI':
                        pros_36 = constants.SI
                        p_pros_36 = 5
                        g_pros_36 = constants.NULO
                    elif ws['H155'].value == 'NO':
                        pros_36 = constants.NO
                        p_pros_36 = 0
                        g_pros_36 = constants.GRAVE
                    if ws['H157'].value == 'SI':
                        pros_37 = constants.SI
                        p_pros_37 = 5
                        g_pros_37 = constants.NULO
                    elif ws['H157'].value == 'NO':
                        pros_37 = constants.NO
                        p_pros_37 = 0
                        g_pros_37 = constants.GRAVE
                    if ws['H158'].value == 'SI':
                        t_pros_38 = constants.ELISA
                        pros_38 = constants.SI
                        p_pros_38 = 5
                        g_pros_38 = constants.NULO
                    elif ws['H159'].value == 'SI':
                        t_pros_38 = constants.QUIMIOLUMINISCENCIA
                        pros_38 = constants.SI
                        p_pros_38 = 5
                        g_pros_38 = constants.NULO
                    elif ws['H160'].value == 'SI':
                        t_pros_38 = constants.ELECTROQUIMIOLUMINISCENCIA
                        pros_38 = constants.SI
                        p_pros_38 = 5
                        g_pros_38 = constants.NULO
                    else:
                        t_pros_38 = False
                        pros_38 = constants.NO
                        p_pros_38 = 0
                        g_pros_38 = constants.GRAVE
                    if ws['H161'].value == 'SI':
                        t_pros_39 = constants.ELISA
                        pros_39 = constants.SI
                        p_pros_39 = 5
                        g_pros_39 = constants.NULO
                    elif ws['H162'].value == 'SI':
                        t_pros_39 = constants.QUIMIOLUMINISCENCIA
                        pros_39 = constants.SI
                        p_pros_39 = 5
                        g_pros_39 = constants.NULO
                    elif ws['H163'].value == 'SI':
                        t_pros_39 = constants.ELECTROQUIMIOLUMINISCENCIA
                        pros_39 = constants.SI
                        p_pros_39 = 5
                        g_pros_39 = constants.NULO
                    else:
                        t_pros_39 = False
                        pros_39 = constants.NO
                        p_pros_39 = 0
                        g_pros_39 = constants.GRAVE
                    if ws['H164'].value == 'SI':
                        t_pros_40 = constants.ELISA
                        pros_40 = constants.SI
                        p_pros_40 = 5
                        g_pros_40 = constants.NULO
                    elif ws['H165'].value == 'SI':
                        t_pros_40 = constants.QUIMIOLUMINISCENCIA
                        pros_40 = constants.SI
                        p_pros_40 = 5
                        g_pros_40 = constants.NULO
                    elif ws['H166'].value == 'SI':
                        t_pros_40 = constants.ELECTROQUIMIOLUMINISCENCIA
                        pros_40 = constants.SI
                        p_pros_40 = 5
                        g_pros_40 = constants.NULO
                    else:
                        t_pros_40 = False
                        pros_40 = constants.NO
                        p_pros_40 = 0
                        g_pros_40 = constants.GRAVE
                    if ws['H167'].value == 'SI':
                        t_pros_41 = constants.ELISA
                        pros_41 = constants.SI
                        p_pros_41 = 5
                        g_pros_41 = constants.NULO
                    elif ws['H168'].value == 'SI':
                        t_pros_41 = constants.QUIMIOLUMINISCENCIA
                        pros_41 = constants.SI
                        p_pros_41 = 5
                        g_pros_41 = constants.NULO
                    elif ws['H169'].value == 'SI':
                        t_pros_41 = constants.ELECTROQUIMIOLUMINISCENCIA
                        pros_41 = constants.SI
                        p_pros_41 = 5
                        g_pros_41 = constants.NULO
                    else:
                        t_pros_41 = False
                        pros_41 = constants.NO
                        p_pros_41 = 0
                        g_pros_41 = constants.GRAVE
                    if ws['H170'].value == 'SI':
                        t_pros_42 = constants.ELISA
                        pros_42 = constants.SI
                        p_pros_42 = 5
                        g_pros_42 = constants.NULO
                    elif ws['H171'].value == 'SI':
                        t_pros_42 = constants.QUIMIOLUMINISCENCIA
                        pros_42 = constants.SI
                        p_pros_42 = 5
                        g_pros_42 = constants.NULO
                    elif ws['H172'].value == 'SI':
                        t_pros_42 = constants.ELECTROQUIMIOLUMINISCENCIA
                        pros_42 = constants.SI
                        p_pros_42 = 5
                        g_pros_42 = constants.NULO
                    else:
                        t_pros_42 = False
                        pros_42 = constants.NO
                        p_pros_42 = 0
                        g_pros_42 = constants.GRAVE
                    if ws['H173'].value == 'SI':
                        t_pros_43 = constants.ELISA
                        pros_43 = constants.SI
                        p_pros_43 = 5
                        g_pros_43 = constants.NULO
                    elif ws['H174'].value == 'SI':
                        t_pros_43 = constants.QUIMIOLUMINISCENCIA
                        pros_43 = constants.SI
                        p_pros_43 = 5
                        g_pros_43 = constants.NULO
                    elif ws['H175'].value == 'SI':
                        t_pros_43 = constants.ELECTROQUIMIOLUMINISCENCIA
                        pros_43 = constants.SI
                        p_pros_43 = 5
                        g_pros_43 = constants.NULO
                    else:
                        t_pros_43 = False
                        pros_43 = constants.NO
                        p_pros_43 = 0
                        g_pros_43 = constants.GRAVE
                    if ws['H176'].value == 'SI':
                        t_pros_44 = constants.ELISA
                        pros_44 = constants.SI
                        p_pros_44 = 5
                        g_pros_44 = constants.NULO
                    elif ws['H177'].value == 'SI':
                        t_pros_44 = constants.QUIMIOLUMINISCENCIA
                        pros_44 = constants.SI
                        p_pros_44 = 5
                        g_pros_44 = constants.NULO
                    elif ws['H178'].value == 'SI':
                        t_pros_44 = constants.ELECTROQUIMIOLUMINISCENCIA
                        pros_44 = constants.SI
                        p_pros_44 = 5
                        g_pros_44 = constants.NULO
                    else:
                        t_pros_44 = False
                        pros_44 = constants.NO
                        p_pros_44 = 0
                        g_pros_44 = constants.GRAVE
                    if ws['H179'].value == 'SI':
                        pros_45 = constants.SI
                    elif ws['H179'].value == 'NO':
                        pros_45 = constants.NO
                    if ws['H180'].value == 'SI':
                        t_pros_47 = constants.DIARIO
                    elif ws['H181'].value == 'SI':
                        t_pros_47 = constants.INTERDIARIO
                    elif ws['H182'].value == 'SI':
                        t_pros_47 = constants.MAYOR_A_UNA_SEMANA
                    else:
                        t_pros_47 = False
                    if ws['H184'].value == 'SI':
                        pros_49 = constants.SI
                        p_pros_49 = 5
                        g_pros_49 = constants.NULO
                    elif ws['H184'].value == 'NO':
                        pros_49 = constants.NO
                        p_pros_49 = 0
                        g_pros_49 = constants.GRAVE
                    if ws['H185'].value == 'SI':
                        pros_50 = constants.SI
                        p_pros_50 = 5
                        g_pros_50 = constants.NULO
                    elif ws['H185'].value == 'NO':
                        pros_50 = constants.NO
                        p_pros_50 = 0
                        g_pros_50 = constants.GRAVE
                    if ws['H186'].value == 'SI':
                        t_pros_50 = constants.NUEVA_MUESTRA
                    elif ws['H187'].value == 'SI':
                        t_pros_50 = constants.CON_MUESTRA_ALMACENADA
                    elif ws['H188'].value == 'SI':
                        t_pros_50 = constants.PLASMA_DE_LA_UNIDAD_EXTRAIDA
                    else:
                        t_pros_50 = False
                    if ws['H189'].value == 'SI':
                        pros_51 = constants.SI
                        p_pros_51 = 5
                        g_pros_51 = constants.NULO
                    elif ws['H189'].value == 'NO':
                        pros_51 = constants.NO
                        p_pros_51 = 0
                        g_pros_51 = constants.GRAVE
                    if ws['H190'].value == 'SI':
                        pros_52 = constants.SI
                        p_pros_52 = 5
                        g_pros_52 = constants.NULO
                    elif ws['H190'].value == 'NO':
                        pros_52 = constants.NO
                        p_pros_52 = 0
                        g_pros_52 = constants.GRAVE
                    if ws['H191'].value == 'SI':
                        pros_53 = constants.SI
                        p_pros_53 = 5
                        g_pros_53 = constants.NULO
                    elif ws['H191'].value == 'NO':
                        pros_53 = constants.NO
                        p_pros_53 = 0
                        g_pros_53 = constants.GRAVE
                    if ws['H192'].value == 'SI':
                        r_pros_53 = constants.SI
                        p_r_pros_53 = 5
                        g_r_pros_53 = constants.NULO
                    elif ws['H192'].value == 'NO':
                        r_pros_53 = constants.NO
                        p_r_pros_53 = 0
                        g_r_pros_53 = constants.GRAVE
                    if ws['H193'].value == 'SI':
                        pros_54 = constants.SI
                        p_pros_54 = 5
                        g_pros_54 = constants.NULO
                    elif ws['H193'].value == 'NO':
                        pros_54 = constants.NO
                        p_pros_54 = 0
                        g_pros_54 = constants.GRAVE
                    if ws['H194'].value == 'SI':
                        pros_55 = constants.SI
                        p_pros_55 = 5
                        g_pros_55 = constants.NULO
                    elif ws['H194'].value == 'NO':
                        pros_55 = constants.NO
                        p_pros_55 = 0
                        g_pros_55 = constants.GRAVE
                    if ws['H196'].value == 'SI':
                        t_pros_56 = constants.MANUAL
                    elif ws['H197'].value == 'SI':
                        t_pros_56 = constants.SEMIAUTOMATICO
                    elif ws['H198'].value == 'SI':
                        t_pros_56 = constants.AUTOMATICO
                    else:
                        t_pros_56 = False
                    if ws['H199'].value == 'SI':
                        pros_57 = constants.SI
                        p_pros_57 = 4
                        g_pros_57 = constants.NULO
                    elif ws['H199'].value == 'NO':
                        pros_57 = constants.NO
                        p_pros_57 = 0
                        g_pros_57 = constants.MODERADO
                    if ws['H200'].value == 'SI':
                        pros_58 = constants.SI
                        p_pros_58 = 4
                        g_pros_58 = constants.NULO
                    elif ws['H200'].value == 'NO':
                        pros_58 = constants.NO
                        p_pros_58 = 0
                        g_pros_58 = constants.MODERADO
                    if ws['H201'].value == 'SI':
                        pros_59 = constants.SI
                        p_pros_59 = 4
                        g_pros_59 = constants.NULO
                    elif ws['H201'].value == 'NO':
                        pros_59 = constants.NO
                        p_pros_59 = 0
                        g_pros_59 = constants.MODERADO
                    if ws['H202'].value == 'SI':
                        pros_60 = constants.SI
                        p_pros_60 = 4
                        g_pros_60 = constants.NULO
                    elif ws['H202'].value == 'NO':
                        pros_60 = constants.NO
                        p_pros_60 = 0
                        g_pros_60 = constants.MODERADO
                    if ws['H203'].value == 'SI':
                        pros_61 = constants.SI
                        p_pros_61 = 4
                        g_pros_61 = constants.NULO
                    elif ws['H203'].value == 'NO':
                        pros_61 = constants.NO
                        p_pros_61 = 0
                        g_pros_61 = constants.MODERADO
                    if ws['H204'].value == 'SI':
                        pros_62 = constants.SI
                        p_pros_62 = 4
                        g_pros_62 = constants.NULO
                    elif ws['H204'].value == 'NO':
                        pros_62 = constants.NO
                        p_pros_62 = 0
                        g_pros_62 = constants.MODERADO
                    if ws['H205'].value == 'SI':
                        pros_63 = constants.SI
                        p_pros_63 = 4
                        g_pros_63 = constants.NULO
                    elif ws['H205'].value == 'NO':
                        pros_63 = constants.NO
                        p_pros_63 = 0
                        g_pros_63 = constants.MODERADO
                    if ws['H206'].value == 'SI':
                        pros_64 = constants.SI
                        p_pros_64 = 4
                        g_pros_64 = constants.NULO
                    elif ws['H206'].value == 'NO':
                        pros_64 = constants.NO
                        p_pros_64 = 0
                        g_pros_64 = constants.MODERADO
                    if ws['H207'].value == 'SI':
                        pros_65 = constants.SI
                        p_pros_65 = 4
                        g_pros_65 = constants.NULO
                    elif ws['H207'].value == 'NO':
                        pros_65 = constants.NO
                        p_pros_65 = 0
                        g_pros_65 = constants.MODERADO
                    if ws['H208'].value == 'SI':
                        pros_66 = constants.SI
                        p_pros_66 = 0
                        g_pros_66 = constants.NULO
                    elif ws['H208'].value == 'NO':
                        pros_66 = constants.NO
                        p_pros_66 = 0
                        g_pros_66 = constants.NULO
                    if ws['H209'].value == 'SI':
                        pros_68 = constants.SI
                    elif ws['H209'].value == 'NO':
                        pros_68 = constants.NO
                    if ws['H210'].value == 'SI':
                        t_pros_68 = constants.PRODUCTOS_IRRADIADOS
                    elif ws['H211'].value == 'SI':
                        t_pros_68 = constants.PRODUCTO_LEUCORREDUCIDO
                    elif ws['H212'].value == 'SI':
                        t_pros_68 = constants.COMPONENTES_EN_POOL
                    else:
                        t_pros_68 = False
                    if ws['H233'].value == 'SI':
                        pros_81 = constants.SI
                    elif ws['H233'].value == 'NO':
                        pros_81 = constants.NO
                    if ws['H234'].value == 'SI':
                        pros_82 = constants.SI
                    elif ws['H234'].value == 'NO':
                        pros_82 = constants.NO
                    if ws['H235'].value == 'SI':
                        pros_83 = constants.SI
                    elif ws['H235'].value == 'NO':
                        pros_83 = constants.NO
                    if ws['H236'].value == 'SI':
                        pros_84 = constants.SI
                    elif ws['H236'].value == 'NO':
                        pros_84 = constants.NO
                    if ws['H238'].value == 'SI':
                        pros_85 = constants.SI
                        p_pros_85 = 4
                        g_pros_85 = constants.NULO
                    elif ws['H238'].value == 'NO':
                        pros_85 = constants.NO
                        p_pros_85 = 0
                        g_pros_85 = constants.MODERADO
                    if ws['H240'].value == 'SI':
                        pros_87 = constants.SI
                    elif ws['H240'].value == 'NO':
                        pros_87 = constants.NO
                    if ws['H242'].value == 'SI':
                        pros_88 = constants.SI
                    elif ws['H242'].value == 'NO':
                        pros_88 = constants.NO
                    if ws['H243'].value == 'SI':
                        pros_89 = constants.SI
                    elif ws['H243'].value == 'NO':
                        pros_89 = constants.NO
                    if ws['H244'].value == 'SI':
                        pros_90 = constants.SI
                    elif ws['H244'].value == 'NO':
                        pros_90 = constants.NO
                    if ws['H245'].value == 'SI':
                        pros_91 = constants.SI
                    elif ws['H245'].value == 'NO':
                        pros_91 = constants.NO
                    if ws['H246'].value == 'SI':
                        pros_92 = constants.SI
                    elif ws['H246'].value == 'NO':
                        pros_92 = constants.NO
                    if ws['H248'].value == 'SI':
                        pros_93 = constants.SI
                        p_pros_93 = 4
                        g_pros_93 = constants.NULO
                    elif ws['H248'].value == 'NO':
                        pros_93 = constants.NO
                        p_pros_93 = 0
                        g_pros_93 = constants.NULO
                    if ws['H249'].value == 'SI':
                        t_pros_93 = constants.METODO
                    elif ws['H250'].value == 'SI':
                        t_pros_93 = constants.EQUIPO
                    else:
                        t_pros_93 = False
                    if ws['H252'].value == 'SI':
                        pros_99 = constants.SI
                        p_pros_99 = 0
                        g_pros_99 = constants.NULO
                    elif ws['H252'].value == 'NO':
                        pros_99 = constants.NO
                        p_pros_99 = 0
                        g_pros_99 = constants.NULO
                    if ws['H253'].value == 'SI':
                        pros_100 = constants.SI
                        p_pros_100 = 0
                        g_pros_100 = constants.NULO
                    elif ws['H253'].value == 'NO':
                        pros_100 = constants.NO
                        p_pros_100 = 0
                        g_pros_100 = constants.NULO
                    if ws['H254'].value == 'SI':
                        pros_101 = constants.SI
                        p_pros_101 = 0
                        g_pros_101 = constants.NULO
                    elif ws['H254'].value == 'NO':
                        pros_101 = constants.NO
                        p_pros_101 = 0
                        g_pros_101 = constants.NULO
                    if ws['H255'].value == 'SI':
                        pros_102 = constants.SI
                        p_pros_102 = 0
                        g_pros_102 = constants.NULO
                    elif ws['H255'].value == 'NO':
                        pros_102 = constants.NO
                        p_pros_102 = 0
                        g_pros_102 = constants.NULO
                    if ws['H256'].value == 'SI':
                        pros_103 = constants.SI
                        p_pros_103 = 0
                        g_pros_103 = constants.NULO
                    elif ws['H256'].value == 'NO':
                        pros_103 = constants.NO
                        p_pros_103 = 0
                        g_pros_103 = constants.NULO
                    if ws['H276'].value == 'SI':
                        cali_4 = constants.SI
                    elif ws['H276'].value == 'NO':
                        cali_4 = constants.NO
                    if ws['H277'].value == 'SI':
                        cali_5 = constants.SI
                    elif ws['H277'].value == 'NO':
                        cali_5 = constants.NO
                    if ws['H278'].value == 'SI':
                        cali_6 = constants.SI
                    elif ws['H278'].value == 'NO':
                        cali_6 = constants.NO
                    if ws['H279'].value == 'SI':
                        cali_7 = constants.SI
                    elif ws['H279'].value == 'NO':
                        cali_7 = constants.NO
                    if ws['H280'].value == 'SI':
                        cali_8 = constants.SI
                    elif ws['H280'].value == 'NO':
                        cali_8 = constants.NO
                    if ws['H311'].value == 'SI':
                        clasf_bio_3 = constants.SI
                        p_clasf_bio_3 = 5
                        g_clasf_bio_3 = constants.NULO
                    elif ws['H311'].value == 'NO':
                        clasf_bio_3 = constants.NO
                        p_clasf_bio_3 = 0
                        g_clasf_bio_3 = constants.GRAVE
                    if ws['H313'].value == 'SI':
                        t_clasf_bio_5 = constants.TECNOLOGO
                    elif ws['H314'].value == 'SI':
                        t_clasf_bio_5 = constants.BIOLOGO
                    elif ws['H315'].value == 'SI':
                        t_clasf_bio_5 = constants.OTRA_PROFESION
                    else:
                        t_clasf_bio_5 = False
                    if ws['H316'].value == 'SI':
                        clasf_bio_6 = constants.SI
                        p_clasf_bio_6 = 5
                        g_clasf_bio_6 = constants.NULO
                    elif ws['H316'].value == 'NO':
                        clasf_bio_6 = constants.NO
                        p_clasf_bio_6 = 0
                        g_clasf_bio_6 = constants.GRAVE
                    if ws['H317'].value == 'SI':
                        clasf_bio_7 = constants.SI
                        p_clasf_bio_7 = 5
                        g_clasf_bio_7 = constants.NULO
                    elif ws['H317'].value == 'NO':
                        clasf_bio_7 = constants.NO
                        p_clasf_bio_7 = 0
                        g_clasf_bio_7 = constants.GRAVE
                    if ws['H328'].value == 'SI':
                        clasf_bio_18 = constants.SI
                        p_clasf_bio_18 = 5
                        g_clasf_bio_18 = constants.NULO
                    elif ws['H328'].value == 'NO':
                        clasf_bio_18 = constants.NO
                        p_clasf_bio_18 = 0
                        g_clasf_bio_18 = constants.GRAVE
                    if ws['H329'].value == 'SI':
                        clasf_bio_19 = constants.SI
                        p_clasf_bio_19 = 5
                        g_clasf_bio_19 = constants.NULO
                    elif ws['H329'].value == 'NO':
                        clasf_bio_19 = constants.NO
                        p_clasf_bio_19 = 0
                        g_clasf_bio_19 = constants.GRAVE
                    if ws['H339'].value == 'SI':
                        equip_8 = constants.SI
                        p_equip_8 = 5
                        g_equip_8 = constants.NULO
                    elif ws['H339'].value == 'NO':
                        equip_8 = constants.NO
                        p_equip_8 = 0
                        g_equip_8 = constants.GRAVE
                    if ws['F351'].value == 'SI':
                        rel_c_equip_1 = constants.SI
                    elif ws['F351'].value == 'NO':
                        rel_c_equip_1 = constants.NO
                    if ws['H351'].value == 'SI':
                        rel_p_equip_1 = constants.SI
                    elif ws['H351'].value == 'NO':
                        rel_p_equip_1 = constants.NO
                    if ws['F356'].value == 'SI':
                        rel_c_equip_2 = constants.SI
                    elif ws['F356'].value == 'NO':
                        rel_c_equip_2 = constants.NO
                    if ws['H356'].value == 'SI':
                        rel_p_equip_2 = constants.SI
                    elif ws['H356'].value == 'NO':
                        rel_p_equip_2 = constants.NO
                    if ws['F376'].value == 'SI':
                        rel_c_equip_6 = constants.SI
                    elif ws['F376'].value == 'NO':
                        rel_c_equip_6 = constants.NO
                    if ws['H376'].value == 'SI':
                        rel_p_equip_6 = constants.SI
                    elif ws['H376'].value == 'NO':
                        rel_p_equip_6 = constants.NO
                    if ws['F391'].value == 'SI':
                        rel_c_equip_9 = constants.SI
                    elif ws['F391'].value == 'NO':
                        rel_c_equip_9 = constants.NO
                    if ws['H391'].value == 'SI':
                        rel_p_equip_9 = constants.SI
                    elif ws['H391'].value == 'NO':
                        rel_p_equip_9 = constants.NO
                    if ws['F396'].value == 'SI':
                        rel_c_equip_10 = constants.SI
                    elif ws['F396'].value == 'NO':
                        rel_c_equip_10 = constants.NO
                    if ws['H396'].value == 'SI':
                        rel_p_equip_10 = constants.SI
                    elif ws['H396'].value == 'NO':
                        rel_p_equip_10 = constants.NO
                    if ws['F401'].value == 'SI':
                        rel_c_equip_11 = constants.SI
                    elif ws['F401'].value == 'NO':
                        rel_c_equip_11 = constants.NO
                    if ws['H401'].value == 'SI':
                        rel_p_equip_11 = constants.SI
                    elif ws['H401'].value == 'NO':
                        rel_p_equip_11 = constants.NO
                    if ws['F411'].value == 'SI':
                        rel_c_equip_13 = constants.SI
                    elif ws['F411'].value == 'NO':
                        rel_c_equip_13 = constants.NO
                    if ws['H411'].value == 'SI':
                        rel_p_equip_13 = constants.SI
                    elif ws['H411'].value == 'NO':
                        rel_p_equip_13 = constants.NO
                    if ws['F466'].value == 'SI':
                        rel_c_equip_24 = constants.SI
                    elif ws['F466'].value == 'NO':
                        rel_c_equip_24 = constants.NO
                    if ws['H466'].value == 'SI':
                        rel_p_equip_24 = constants.SI
                    elif ws['H466'].value == 'NO':
                        rel_p_equip_24 = constants.NO
                    if ws['F471'].value == 'SI':
                        rel_c_equip_25 = constants.SI
                    elif ws['F471'].value == 'NO':
                        rel_c_equip_25 = constants.NO
                    if ws['H471'].value == 'SI':
                        rel_p_equip_25 = constants.SI
                    elif ws['H471'].value == 'NO':
                        rel_p_equip_25 = constants.NO
                    if ws['F476'].value == 'SI':
                        rel_c_equip_26 = constants.SI
                    elif ws['F476'].value == 'NO':
                        rel_c_equip_26 = constants.NO
                    if ws['H476'].value == 'SI':
                        rel_p_equip_26 = constants.SI
                    elif ws['H476'].value == 'NO':
                        rel_p_equip_26 = constants.NO
                    if ws['F491'].value == 'SI':
                        rel_c_equip_33 = constants.SI
                    elif ws['F491'].value == 'NO':
                        rel_c_equip_33 = constants.NO
                    if ws['H491'].value == 'SI':
                        rel_p_equip_33 = constants.SI
                    elif ws['H491'].value == 'NO':
                        rel_p_equip_33 = constants.NO
                    if ws['F496'].value == 'SI':
                        rel_c_equip_34 = constants.SI
                    elif ws['F496'].value == 'NO':
                        rel_c_equip_34 = constants.NO
                    if ws['H496'].value == 'SI':
                        rel_p_equip_34 = constants.SI
                    elif ws['H496'].value == 'NO':
                        rel_p_equip_34 = constants.NO
                    if ws['H534'].value == 'SI':
                        infra_instal_5 = constants.SI
                        p_infra_instal_5 = 5
                        g_infra_instal_5 = constants.NULO
                    elif ws['H534'].value == 'NO':
                        infra_instal_5 = constants.NO
                        p_infra_instal_5 = 0
                        g_infra_instal_5 = constants.MODERADO
                    if ws['H536'].value == 'SI':
                        infra_instal_7 = constants.SI
                        p_infra_instal_7 = 5
                        g_infra_instal_7 = constants.NULO
                    elif ws['H536'].value == 'NO':
                        infra_instal_7 = constants.NO
                        p_infra_instal_7 = 0
                        g_infra_instal_7 = constants.MODERADO
                    if ws['H537'].value == 'SI':
                        infra_instal_8 = constants.SI
                        p_infra_instal_8 = 5
                        g_infra_instal_8 = constants.NULO
                    elif ws['H537'].value == 'NO':
                        infra_instal_8 = constants.NO
                        p_infra_instal_8 = 0
                        g_infra_instal_8 = constants.MODERADO
                    if ws['H538'].value == 'SI':
                        infra_instal_9 = constants.SI
                        p_infra_instal_9 = 5
                        g_infra_instal_9 = constants.NULO
                    elif ws['H538'].value == 'NO':
                        infra_instal_9 = constants.NO
                        p_infra_instal_9 = 0
                        g_infra_instal_9 = constants.MODERADO
                    if ws['H539'].value == 'SI':
                        infra_instal_10 = constants.SI
                    elif ws['H539'].value == 'NO':
                        infra_instal_10 = constants.NO
                    if ws['H541'].value == 'SI':
                        infra_instal_12 = constants.SI
                        p_infra_instal_12 = 5
                        g_infra_instal_12 = constants.NULO
                    elif ws['H541'].value == 'NO':
                        infra_instal_12 = constants.NO
                        p_infra_instal_12 = 0
                        g_infra_instal_12 = constants.MODERADO
                    if ws['H542'].value == 'SI':
                        infra_instal_13 = constants.SI
                        p_infra_instal_13 = 5
                        g_infra_instal_13 = constants.NULO
                    elif ws['H542'].value == 'NO':
                        infra_instal_13 = constants.NO
                        p_infra_instal_13 = 0
                        g_infra_instal_13 = constants.MODERADO
                    if ws['H544'].value == 'SI':
                        infra_instal_15 = constants.SI
                        p_infra_instal_15 = 5
                        g_infra_instal_15 = constants.NULO
                    elif ws['H544'].value == 'NO':
                        infra_instal_15 = constants.NO
                        p_infra_instal_15 = 0
                        g_infra_instal_15 = constants.MODERADO
                    if ws['H545'].value == 'SI':
                        infra_instal_16 = constants.SI
                        p_infra_instal_16 = 5
                        g_infra_instal_16 = constants.NULO
                    elif ws['H545'].value == 'NO':
                        infra_instal_16 = constants.NO
                        p_infra_instal_16 = 0
                        g_infra_instal_16 = constants.MODERADO
                if ws['H46'].value == 'SI':
                    rg_1 = constants.SI
                    p_rg_1 = 5
                    g_rg_1 = constants.NULO
                elif ws['H46'].value == 'NO':
                    rg_1 = constants.NO
                    p_rg_1 = 0
                    g_rg_1 = constants.LEVE
                if ws['H47'].value == 'SI':
                    rg_2 = constants.SI
                    p_rg_2 = 5
                    g_rg_2 = constants.NULO
                elif ws['H47'].value == 'NO':
                    rg_2 = constants.NO
                    p_rg_2 = 0
                    g_rg_2 = constants.LEVE
                if ws['H48'].value == 'SI':
                    t_rg_3 = constants.SERVICIO
                    p_rg_3 = 0
                elif ws['H49'].value == 'SI':
                    t_rg_3 = constants.DEPARTAMENTO
                    p_rg_3 = 2
                elif ws['H50'].value == 'SI':
                    t_rg_3 = constants.AREA
                    p_rg_3 = 3
                else:
                    t_rg_3 = False
                    p_rg_3 = 0
                if ws['H51'].value == 'SI':
                    rg_4 = constants.SI
                    p_rg_4 = 4
                    g_rg_4 = constants.NULO
                elif ws['H51'].value == 'NO':
                    rg_4 = constants.NO
                    p_rg_4 = 0
                    g_rg_4 = constants.LEVE
                if ws['H52'].value == 'SI':
                    rg_5 = constants.SI
                    p_rg_5 = 4
                    g_rg_5 = constants.NULO
                elif ws['H52'].value == 'NO':
                    rg_5 = constants.NO
                    p_rg_5 = 0
                    g_rg_5 = constants.LEVE
                if ws['H54'].value == 'SI':
                    rg_8 = constants.SI
                    p_rg_8 = 5
                    g_rg_8 = constants.NULO
                elif ws['H54'].value == 'NO':
                    rg_8 = constants.NO
                    p_rg_8 = 0
                    g_rg_8 = constants.LEVE
                if ws['H55'].value == 'SI':
                    rg_9 = constants.SI
                    p_rg_9 = 5
                    g_rg_9 = constants.NULO
                elif ws['H55'].value == 'NO':
                    rg_9 = constants.NO
                    p_rg_9 = 0
                    g_rg_9 = constants.LEVE
                if ws['H56'].value == 'SI':
                    rg_10 = constants.SI
                    p_rg_10 = 5
                    g_rg_10 = constants.NULO
                elif ws['H56'].value == 'NO':
                    rg_10 = constants.NO
                    p_rg_10 = 0
                    g_rg_10 = constants.LEVE
                if ws['H57'].value == 'SI':
                    rg_11 = constants.SI
                    p_rg_11 = 5
                    g_rg_11 = constants.NULO
                elif ws['H57'].value == 'NO':
                    rg_11 = constants.NO
                    p_rg_11 = 0
                    g_rg_11 = constants.LEVE
                if ws['H58'].value == 'SI':
                    rg_12 = constants.SI
                    p_rg_12 = 5
                    g_rg_12 = constants.NULO
                elif ws['H58'].value == 'NO':
                    rg_12 = constants.NO
                    p_rg_12 = 0
                    g_rg_12 = constants.LEVE
                if ws['H59'].value == 'SI':
                    rg_13 = constants.SI
                    p_rg_13 = 5
                    g_rg_13 = constants.NULO
                elif ws['H59'].value == 'NO':
                    rg_13 = constants.NO
                    p_rg_13 = 0
                    g_rg_13 = constants.LEVE
                if ws['H61'].value == 'SI':
                    rg_15 = constants.SI
                    p_rg_15 = 5
                    g_rg_15 = constants.NULO
                elif ws['H61'].value == 'NO':
                    rg_15 = constants.NO
                    p_rg_15 = 0
                    g_rg_15 = constants.LEVE
                if ws['H62'].value == 'SI':
                    rg_16 = constants.SI
                    p_rg_16 = 4
                    g_rg_16 = constants.NULO
                elif ws['H62'].value == 'NO':
                    rg_16 = constants.NO
                    p_rg_16 = 0
                    g_rg_16 = constants.LEVE
                if ws['H64'].value == 'SI':
                    rg_18 = constants.SI
                    p_rg_18 = 4
                    g_rg_18 = constants.NULO
                elif ws['H64'].value == 'NO':
                    rg_18 = constants.NO
                    p_rg_18 = 0
                    g_rg_18 = constants.LEVE
                if ws['H66'].value == 'SI':
                    rg_19 = constants.SI
                    p_rg_19 = 5
                    g_rg_19 = constants.NULO
                elif ws['H66'].value == 'NO':
                    rg_19 = constants.NO
                    p_rg_19 = 0
                    g_rg_19 = constants.GRAVE
                if ws['H67'].value == 'SI':
                    t_rg_20 = constants.NOMBRADO
                    p_rg_20 = 4
                    g_rg_20 = constants.NULO
                elif ws['H68'].value == 'SI':
                    t_rg_20 = constants.CAS
                    p_rg_20 = 4
                    g_rg_20 = constants.NULO
                elif ws['H69'].value == 'SI':
                    t_rg_20 = constants.TERCERO
                    p_rg_20 = 4
                    g_rg_20 = constants.LEVE
                else:
                    t_rg_20 = False
                    p_rg_20 = 0
                    g_rg_20 = constants.NULO
                if ws['H71'].value == 'SI':
                    rg_21 = constants.SI
                    p_rg_21 = 3
                    g_rg_21 = constants.NULO
                elif ws['H71'].value == 'NO':
                    rg_21 = constants.NO
                    p_rg_21 = 0
                    g_rg_21 = constants.NULO
                if ws['H78'].value == 'SI':
                    rg_23 = constants.SI
                    p_rg_23 = 4
                    g_rg_23 = constants.NULO
                elif ws['H78'].value == 'NO':
                    rg_23 = constants.NO
                    p_rg_23 = 0
                    g_rg_23 = constants.NULO
                if ws['H79'].value == 'SI':
                    rg_24 = constants.SI
                    p_rg_24 = 4
                    g_rg_24 = constants.NULO
                elif ws['H79'].value == 'NO':
                    rg_24 = constants.NO
                    p_rg_24 = 0
                    g_rg_24 = constants.MODERADO
                if ws['H80'].value == 'SI':
                    rg_25 = constants.SI
                    p_rg_25 = 5
                    g_rg_25 = constants.NULO
                elif ws['H80'].value == 'NO':
                    rg_25 = constants.NO
                    p_rg_25 = 0
                    g_rg_25 = constants.GRAVE
                if ws['H82'].value == 'SI':
                    rg_28 = constants.SI
                    p_rg_28 = 5
                    g_rg_28 = constants.NULO
                elif ws['H82'].value == 'NO':
                    rg_28 = constants.NO
                    p_rg_28 = 0
                    g_rg_28 = constants.GRAVE
                if ws['H83'].value == 'SI':
                    rg_29 = constants.SI
                    p_rg_29 = 5
                    g_rg_29 = constants.NULO
                elif ws['H83'].value == 'NO':
                    rg_29 = constants.NO
                    p_rg_29 = 0
                    g_rg_29 = constants.GRAVE
                if ws['H84'].value == 'SI':
                    rg_30 = constants.SI
                    p_rg_30 = 5
                    g_rg_30 = constants.NULO
                elif ws['H84'].value == 'NO':
                    rg_30 = constants.NO
                    p_rg_30 = 0
                    g_rg_30 = constants.LEVE
                if ws['H88'].value == 'SI':
                    rh_1 = constants.SI
                    p_rh_1 = 5
                    g_rh_1 = constants.NULO
                elif ws['H88'].value == 'NO':
                    rh_1 = constants.NO
                    p_rh_1 = 0
                    g_rh_1 = constants.GRAVE
                if ws['H89'].value == 'SI':
                    rh_2 = constants.SI
                    p_rh_2 = 4
                    g_rh_2 = constants.NULO
                elif ws['H89'].value == 'NO':
                    rh_2 = constants.NO
                    p_rh_2 = 0
                    g_rh_2 = constants.NULO
                if ws['H90'].value == 'SI':
                    rh_3 = constants.SI
                    p_rh_3 = 5
                    g_rh_3 = constants.NULO
                elif ws['H90'].value == 'NO':
                    rh_3 = constants.NO
                    p_rh_3 = 0
                    g_rh_3 = constants.GRAVE
                if ws['H91'].value == 'SI':
                    rh_4 = constants.SI
                    p_rh_4 = 4
                    g_rh_4 = constants.NULO
                elif ws['H91'].value == 'NO':
                    rh_4 = constants.NO
                    p_rh_4 = 0
                    g_rh_4 = constants.LEVE
                if ws['H92'].value == 'SI':
                    rh_5 = constants.SI
                    p_rh_5 = 4
                    g_rh_5 = constants.NULO
                elif ws['H92'].value == 'NO':
                    rh_5 = constants.NO
                    p_rh_5 = 0
                    g_rh_5 = constants.GRAVE
                if ws['H93'].value == 'SI':
                    rh_6 = constants.SI
                    p_rh_6 = 2
                    g_rh_6 = constants.NULO
                elif ws['H93'].value == 'NO':
                    rh_6 = constants.NO
                    p_rh_6 = 0
                    g_rh_6 = constants.NULO
                if ws['H94'].value == 'SI':
                    rh_7 = constants.SI
                    p_rh_7 = 1
                    g_rh_7 = constants.NULO
                elif ws['H94'].value == 'NO':
                    rh_7 = constants.NO
                    p_rh_7 = 0
                    g_rh_7 = constants.NULO
                if ws['H95'].value == 'SI':
                    rh_8 = constants.SI
                    p_rh_8 = 3
                    g_rh_8 = constants.NULO
                elif ws['H95'].value == 'NO':
                    rh_8 = constants.NO
                    p_rh_8 = 0
                    g_rh_8 = constants.NULO
                if ws['H96'].value == 'SI':
                    rh_9 = constants.SI
                    p_rh_9 = 3
                    g_rh_9 = constants.NULO
                elif ws['H96'].value == 'NO':
                    rh_9 = constants.NO
                    p_rh_9 = 0
                    g_rh_9 = constants.LEVE
                if ws['H97'].value == 'SI':
                    rh_10 = constants.SI
                    p_rh_10 = 5
                    g_rh_10 = constants.NULO
                elif ws['H97'].value == 'NO':
                    rh_10 = constants.NO
                    p_rh_10 = 0
                    g_rh_10 = constants.GRAVE
                if ws['H151'].value == 'SI':
                    pros_34 = constants.SI
                    p_pros_34 = 4
                    g_pros_34 = constants.NULO
                elif ws['H151'].value == 'NO':
                    pros_34 = constants.NO
                    p_pros_34 = 0
                    g_pros_34 = constants.MODERADO
                if ws['H152'].value == 'SI':
                    t_pros_34 = constants.MANUAL
                elif ws['H153'].value == 'SI':
                    t_pros_34 = constants.DIGITAL
                else:
                    t_pros_34 = False
                if ws['H215'].value == 'SI':
                    pros_69 = constants.SI
                    p_pros_69 = 5
                    g_pros_69 = constants.NULO
                elif ws['H215'].value == 'NO':
                    pros_69 = constants.NO
                    p_pros_69 = 0
                    g_pros_69 = constants.GRAVE
                if ws['H216'].value == 'SI':
                    pros_70 = constants.SI
                    p_pros_70 = 5
                    g_pros_70 = constants.NULO
                elif ws['H216'].value == 'NO':
                    pros_70 = constants.NO
                    p_pros_70 = 0
                    g_pros_70 = constants.GRAVE
                if ws['H218'].value == 'SI':
                    pros_73 = constants.SI
                    p_pros_73 = 5
                    g_pros_73 = constants.NULO
                elif ws['H218'].value == 'NO':
                    pros_73 = constants.NO
                    p_pros_73 = 0
                    g_pros_73 = constants.GRAVE
                if ws['H219'].value == 'SI':
                    pros_74 = constants.SI
                    p_pros_74 = 5
                    g_pros_74 = constants.NULO
                elif ws['H219'].value == 'NO':
                    pros_74 = constants.NO
                    p_pros_74 = 0
                    g_pros_74 = constants.GRAVE
                if ws['H220'].value == 'SI':
                    pros_75 = constants.SI
                    p_pros_75 = 5
                    g_pros_75 = constants.NULO
                elif ws['H220'].value == 'NO':
                    pros_75 = constants.NO
                    p_pros_75 = 0
                    g_pros_75 = constants.GRAVE
                if ws['H221'].value == 'SI':
                    pros_76 = constants.SI
                    p_pros_76 = 4
                    g_pros_76 = constants.NULO
                elif ws['H221'].value == 'NO':
                    pros_76 = constants.NO
                    p_pros_76 = 0
                    g_pros_76 = constants.MODERADO
                if ws['H222'].value == 'SI':
                    pros_77 = constants.SI
                    p_pros_77 = 5
                    g_pros_77 = constants.NULO
                elif ws['H222'].value == 'NO':
                    pros_77 = constants.NO
                    p_pros_77 = 0
                    g_pros_77 = constants.GRAVE
                if ws['H223'].value == 'SI':
                    t_pros_77 = constants.MANUAL
                elif ws['H224'].value == 'SI':
                    t_pros_77 = constants.SEMIAUTOMATICO
                elif ws['H225'].value == 'SI':
                    t_pros_77 = constants.AUTOMATICO
                else:
                    t_pros_77 = False
                if ws['H227'].value == 'SI':
                    pros_78 = constants.SI
                    p_pros_78 = 5
                    g_pros_78 = constants.NULO
                elif ws['H227'].value == 'NO':
                    pros_78 = constants.NO
                    p_pros_78 = 0
                    g_pros_78 = constants.MODERADO
                if ws['H229'].value == 'SI':
                    pros_79 = constants.SI
                    p_pros_79 = 4
                    g_pros_79 = constants.NULO
                elif ws['H229'].value == 'NO':
                    pros_79 = constants.NO
                    p_pros_79 = 0
                    g_pros_79 = constants.MODERADO
                if ws['H230'].value == 'SI':
                    pros_80 = constants.SI
                    p_pros_80 = 5
                    g_pros_80 = constants.NULO
                elif ws['H230'].value == 'NO':
                    pros_80 = constants.NO
                    p_pros_80 = 0
                    g_pros_80 = constants.GRAVE
                if ws['H258'].value == 'SI':
                    pros_94 = constants.SI
                    p_pros_94 = 5
                    g_pros_94 = constants.NULO
                elif ws['H258'].value == 'NO':
                    pros_94 = constants.NO
                    p_pros_94 = 0
                    g_pros_94 = constants.MODERADO
                if ws['H259'].value == 'SI':
                    pros_95 = constants.SI
                    p_pros_95 = 5
                    g_pros_95 = constants.NULO
                elif ws['H259'].value == 'NO':
                    pros_95 = constants.NO
                    p_pros_95 = 0
                    g_pros_95 = constants.MODERADO
                if ws['H260'].value == 'SI':
                    pros_96 = constants.SI
                elif ws['H260'].value == 'NO':
                    pros_96 = constants.NO
                if ws['H261'].value == 'SI':
                    pros_97 = constants.SI
                elif ws['H261'].value == 'NO':
                    pros_97 = constants.NO
                if ws['H262'].value == 'SI':
                    pros_98 = constants.SI
                    p_pros_98 = 5
                    g_pros_98 = constants.NULO
                elif ws['H262'].value == 'NO':
                    pros_98 = constants.NO
                    p_pros_98 = 0
                    g_pros_98 = constants.MODERADO
                if ws['H268'].value == 'SI':
                    cali_1 = constants.SI
                elif ws['H268'].value == 'NO':
                    cali_1 = constants.NO
                if ws['H269'].value == 'SI':
                    cali_2 = constants.SI
                elif ws['H269'].value == 'NO':
                    cali_2 = constants.NO
                if ws['H170'].value == 'SI':
                    t_cali_2 = constants.PEVED
                elif ws['H171'].value == 'SI':
                    t_cali_2 = constants.SO
                elif ws['H172'].value == 'SI':
                    t_cali_2 = constants.TO
                elif ws['H173'].value == 'SI':
                    t_cali_2 = constants.OTRO
                else:
                    t_cali_2 = False
                if ws['H274'].value == 'SI':
                    cali_3 = constants.SI
                elif ws['H274'].value == 'NO':
                    cali_3 = constants.NO
                if ws['H289'].value == 'SI':
                    alm_6 = constants.SI
                    p_alm_6 = 5
                    g_alm_6 = constants.NULO
                elif ws['H289'].value == 'NO':
                    alm_6 = constants.NO
                    p_alm_6 = 0
                    g_alm_6 = constants.GRAVE
                if ws['H290'].value == 'SI':
                    alm_7 = constants.SI
                    p_alm_7 = 4
                    g_alm_7 = constants.NULO
                elif ws['H290'].value == 'NO':
                    alm_7 = constants.NO
                    p_alm_7 = 0
                    g_alm_7 = constants.MODERADO
                if ws['H291'].value == 'SI':
                    alm_8 = constants.SI
                    p_alm_8 = 5
                    g_alm_8 = constants.NULO
                elif ws['H291'].value == 'NO':
                    alm_8 = constants.NO
                    p_alm_8 = 0
                    g_alm_8 = constants.GRAVE
                if ws['H301'].value == 'SI':
                    sis_inform_1 = constants.SI
                elif ws['H301'].value == 'NO':
                    sis_inform_1 = constants.NO
                if ws['H304'].value == 'SI':
                    t_sis_inform_4 = constants.PROPIO
                elif ws['H305'].value == 'SI':
                    t_sis_inform_4 = constants.CESION_DE_USO
                else:
                    t_sis_inform_4 = False
                if ws['H306'].value == 'SI':
                    sis_inform_5 = constants.SI
                elif ws['H306'].value == 'NO':
                    sis_inform_5 = constants.NO
                if ws['H310'].value == 'SI':
                    clasf_bio_2 = constants.SI
                    p_clasf_bio_2 = 5
                    g_clasf_bio_2 = constants.NULO
                elif ws['H310'].value == 'NO':
                    clasf_bio_2 = constants.NO
                    p_clasf_bio_2 = 0
                    g_clasf_bio_2 = constants.GRAVE
                if ws['H318'].value == 'SI':
                    clasf_bio_8 = constants.SI
                    p_clasf_bio_8 = 5
                    g_clasf_bio_8 = constants.NULO
                elif ws['H318'].value == 'NO':
                    clasf_bio_8 = constants.NO
                    p_clasf_bio_8 = 0
                    g_clasf_bio_8 = constants.GRAVE
                if ws['H319'].value == 'SI':
                    t_clasf_bio_9 = constants.MENOR_UN_ANO
                    p_clasf_bio_9 = 3
                    g_clasf_bio_9 = constants.GRAVE
                elif ws['H320'].value == 'SI':
                    t_clasf_bio_9 = constants.DE_UNO_A_DOS
                    p_clasf_bio_9 = 4
                    g_clasf_bio_9 = constants.MODERADO
                elif ws['H321'].value == 'SI':
                    t_clasf_bio_9 = constants.DE_DOS_A_TRES
                    p_clasf_bio_9 = 4
                    g_clasf_bio_9 = constants.NULO
                elif ws['H322'].value == 'SI':
                    t_clasf_bio_9 = constants.DE_TRES_A_CUATRO
                    p_clasf_bio_9 = 5
                    g_clasf_bio_9 = constants.NULO
                elif ws['H323'].value == 'SI':
                    t_clasf_bio_9 = constants.DE_CINCO_A_MAS
                    p_clasf_bio_9 = 5
                    g_clasf_bio_9 = constants.NULO
                else:
                    t_clasf_bio_9 = False
                    p_clasf_bio_9 = 0
                    g_clasf_bio_9 = False
                if ws['H325'].value == 'SI':
                    clasf_bio_14 = constants.SI
                    p_clasf_bio_14 = 5
                    g_clasf_bio_14 = constants.NULO
                elif ws['H325'].value == 'NO':
                    clasf_bio_14 = constants.NO
                    p_clasf_bio_14 = 0
                    g_clasf_bio_14 = constants.GRAVE
                if ws['H326'].value == 'SI':
                    clasf_bio_15 = constants.SI
                    p_clasf_bio_15 = 5
                    g_clasf_bio_15 = constants.NULO
                elif ws['H326'].value == 'NO':
                    clasf_bio_15 = constants.NO
                    p_clasf_bio_15 = 0
                    g_clasf_bio_15 = constants.GRAVE
                if ws['H332'].value == 'SI':
                    equip_1 = constants.SI
                    p_equip_1 = 5
                    g_equip_1 = constants.NULO
                elif ws['H332'].value == 'NO':
                    equip_1 = constants.NO
                    p_equip_1 = 0
                    g_equip_1 = constants.GRAVE
                if ws['H333'].value == 'SI':
                    equip_2 = constants.SI
                    p_equip_2 = 5
                    g_equip_2 = constants.NULO
                elif ws['H333'].value == 'NO':
                    equip_2 = constants.NO
                    p_equip_2 = 0
                    g_equip_2 = constants.GRAVE
                if ws['H334'].value == 'SI':
                    equip_3 = constants.SI
                    p_equip_3 = 4
                    g_equip_3 = constants.NULO
                elif ws['H334'].value == 'NO':
                    equip_3 = constants.NO
                    p_equip_3 = 0
                    g_equip_3 = constants.MODERADO
                if ws['H335'].value == 'SI':
                    equip_4 = constants.SI
                    p_equip_4 = 3
                    g_equip_4 = constants.NULO
                elif ws['H335'].value == 'NO':
                    equip_4 = constants.NO
                    p_equip_4 = 0
                    g_equip_4 = constants.MODERADO
                if ws['H336'].value == 'SI':
                    equip_5 = constants.SI
                    p_equip_5 = 5
                    g_equip_5 = constants.NULO
                elif ws['H336'].value == 'NO':
                    equip_5 = constants.NO
                    p_equip_5 = 0
                    g_equip_5 = constants.GRAVE
                if ws['H337'].value == 'SI':
                    equip_6 = constants.SI
                    p_equip_6 = 5
                    g_equip_6 = constants.NULO
                elif ws['H337'].value == 'NO':
                    equip_6 = constants.NO
                    p_equip_6 = 0
                    g_equip_6 = constants.GRAVE
                if ws['H338'].value == 'SI':
                    equip_7 = constants.SI
                    p_equip_7 = 4
                    g_equip_7 = constants.NULO
                elif ws['H338'].value == 'NO':
                    equip_7 = constants.NO
                    p_equip_7 = 0
                    g_equip_7 = constants.MODERADO
                if ws['H340'].value == 'SI':
                    equip_9 = constants.SI
                    p_equip_9 = 5
                    g_equip_9 = constants.NULO
                elif ws['H340'].value == 'NO':
                    equip_9 = constants.NO
                    p_equip_9 = 0
                    g_equip_9 = constants.GRAVE
                if ws['H341'].value == 'SI':
                    equip_10 = constants.SI
                    p_equip_10 = 5
                    g_equip_10 = constants.NULO
                elif ws['H341'].value == 'NO':
                    equip_10 = constants.NO
                    p_equip_10 = 0
                    g_equip_10 = constants.MODERADO
                if ws['H342'].value == 'SI':
                    equip_11 = constants.SI
                    p_equip_11 = 5
                    g_equip_11 = constants.NULO
                elif ws['H342'].value == 'NO':
                    equip_11 = constants.NO
                    p_equip_11 = 0
                    g_equip_11 = constants.MODERADO
                if ws['F361'].value == 'SI':
                    rel_c_equip_3 = constants.SI
                elif ws['F361'].value == 'NO':
                    rel_c_equip_3 = constants.NO
                if ws['H361'].value == 'SI':
                    rel_p_equip_3 = constants.SI
                elif ws['H361'].value == 'NO':
                    rel_p_equip_3 = constants.NO
                if ws['F366'].value == 'SI':
                    rel_c_equip_4 = constants.SI
                elif ws['F366'].value == 'NO':
                    rel_c_equip_4 = constants.NO
                if ws['H366'].value == 'SI':
                    rel_p_equip_4 = constants.SI
                elif ws['H366'].value == 'NO':
                    rel_p_equip_4 = constants.NO
                if ws['F371'].value == 'SI':
                    rel_c_equip_5 = constants.SI
                elif ws['F371'].value == 'NO':
                    rel_c_equip_5 = constants.NO
                if ws['H371'].value == 'SI':
                    rel_p_equip_5 = constants.SI
                elif ws['H371'].value == 'NO':
                    rel_p_equip_5 = constants.NO
                if ws['F381'].value == 'SI':
                    rel_c_equip_7 = constants.SI
                elif ws['F381'].value == 'NO':
                    rel_c_equip_7 = constants.NO
                if ws['H381'].value == 'SI':
                    rel_p_equip_7 = constants.SI
                elif ws['H381'].value == 'NO':
                    rel_p_equip_7 = constants.NO
                if ws['F386'].value == 'SI':
                    rel_c_equip_8 = constants.SI
                elif ws['F386'].value == 'NO':
                    rel_c_equip_8 = constants.NO
                if ws['H386'].value == 'SI':
                    rel_p_equip_8 = constants.SI
                elif ws['H386'].value == 'NO':
                    rel_p_equip_8 = constants.NO
                if ws['F406'].value == 'SI':
                    rel_c_equip_12 = constants.SI
                elif ws['F406'].value == 'NO':
                    rel_c_equip_12 = constants.NO
                if ws['H406'].value == 'SI':
                    rel_p_equip_12 = constants.SI
                elif ws['H406'].value == 'NO':
                    rel_p_equip_12 = constants.NO
                if ws['F416'].value == 'SI':
                    rel_c_equip_14 = constants.SI
                elif ws['F416'].value == 'NO':
                    rel_c_equip_14 = constants.NO
                if ws['H416'].value == 'SI':
                    rel_p_equip_14 = constants.SI
                elif ws['H416'].value == 'NO':
                    rel_p_equip_14 = constants.NO
                if ws['F421'].value == 'SI':
                    rel_c_equip_15 = constants.SI
                elif ws['F421'].value == 'NO':
                    rel_c_equip_15 = constants.NO
                if ws['H421'].value == 'SI':
                    rel_p_equip_15 = constants.SI
                elif ws['H421'].value == 'NO':
                    rel_p_equip_15 = constants.NO
                if ws['F426'].value == 'SI':
                    rel_c_equip_16 = constants.SI
                elif ws['F426'].value == 'NO':
                    rel_c_equip_16 = constants.NO
                if ws['H426'].value == 'SI':
                    rel_p_equip_16 = constants.SI
                elif ws['H426'].value == 'NO':
                    rel_p_equip_16 = constants.NO
                if ws['F431'].value == 'SI':
                    rel_c_equip_17 = constants.SI
                elif ws['F431'].value == 'NO':
                    rel_c_equip_17 = constants.NO
                if ws['H431'].value == 'SI':
                    rel_p_equip_17 = constants.SI
                elif ws['H431'].value == 'NO':
                    rel_p_equip_17 = constants.NO
                if ws['F436'].value == 'SI':
                    rel_c_equip_18 = constants.SI
                elif ws['F436'].value == 'NO':
                    rel_c_equip_18 = constants.NO
                if ws['H436'].value == 'SI':
                    rel_p_equip_18 = constants.SI
                elif ws['H436'].value == 'NO':
                    rel_p_equip_18 = constants.NO
                if ws['F441'].value == 'SI':
                    rel_c_equip_19 = constants.SI
                elif ws['F441'].value == 'NO':
                    rel_c_equip_19 = constants.NO
                if ws['H441'].value == 'SI':
                    rel_p_equip_19 = constants.SI
                elif ws['H441'].value == 'NO':
                    rel_p_equip_19 = constants.NO
                if ws['F446'].value == 'SI':
                    rel_c_equip_20 = constants.SI
                elif ws['F446'].value == 'NO':
                    rel_c_equip_20 = constants.NO
                if ws['H446'].value == 'SI':
                    rel_p_equip_20 = constants.SI
                elif ws['H446'].value == 'NO':
                    rel_p_equip_20 = constants.NO
                if ws['F451'].value == 'SI':
                    rel_c_equip_21 = constants.SI
                elif ws['F451'].value == 'NO':
                    rel_c_equip_21 = constants.NO
                if ws['H451'].value == 'SI':
                    rel_p_equip_21 = constants.SI
                elif ws['H451'].value == 'NO':
                    rel_p_equip_21 = constants.NO
                if ws['F456'].value == 'SI':
                    rel_c_equip_22 = constants.SI
                elif ws['F456'].value == 'NO':
                    rel_c_equip_22 = constants.NO
                if ws['H456'].value == 'SI':
                    rel_p_equip_22 = constants.SI
                elif ws['H456'].value == 'NO':
                    rel_p_equip_22 = constants.NO
                if ws['F461'].value == 'SI':
                    rel_c_equip_23 = constants.SI
                elif ws['F461'].value == 'NO':
                    rel_c_equip_23 = constants.NO
                if ws['H461'].value == 'SI':
                    rel_p_equip_23 = constants.SI
                elif ws['H461'].value == 'NO':
                    rel_p_equip_23 = constants.NO
                if ws['F481'].value == 'SI':
                    rel_c_equip_31 = constants.SI
                elif ws['F481'].value == 'NO':
                    rel_c_equip_31 = constants.NO
                if ws['H481'].value == 'SI':
                    rel_p_equip_31 = constants.SI
                elif ws['H481'].value == 'NO':
                    rel_p_equip_31 = constants.NO
                if ws['F486'].value == 'SI':
                    rel_c_equip_32 = constants.SI
                elif ws['F486'].value == 'NO':
                    rel_c_equip_32 = constants.NO
                if ws['H486'].value == 'SI':
                    rel_p_equip_32 = constants.SI
                elif ws['H486'].value == 'NO':
                    rel_p_equip_32 = constants.NO
                if ws['F504'].value == 'SI':
                    rel_c_equip_27 = constants.SI
                elif ws['F504'].value == 'NO':
                    rel_c_equip_27 = constants.NO
                if ws['H504'].value == 'SI':
                    rel_p_equip_27 = constants.SI
                elif ws['H504'].value == 'NO':
                    rel_p_equip_27 = constants.NO
                if ws['F509'].value == 'SI':
                    rel_c_equip_28 = constants.SI
                elif ws['F509'].value == 'NO':
                    rel_c_equip_28 = constants.NO
                if ws['H509'].value == 'SI':
                    rel_p_equip_28 = constants.SI
                elif ws['H509'].value == 'NO':
                    rel_p_equip_28 = constants.NO
                if ws['F514'].value == 'SI':
                    rel_c_equip_29 = constants.SI
                elif ws['F514'].value == 'NO':
                    rel_c_equip_29 = constants.NO
                if ws['H514'].value == 'SI':
                    rel_p_equip_29 = constants.SI
                elif ws['H514'].value == 'NO':
                    rel_p_equip_29 = constants.NO
                if ws['F519'].value == 'SI':
                    rel_c_equip_30 = constants.SI
                elif ws['F519'].value == 'NO':
                    rel_c_equip_30 = constants.NO
                if ws['H519'].value == 'SI':
                    rel_p_equip_30 = constants.SI
                elif ws['H519'].value == 'NO':
                    rel_p_equip_30 = constants.NO
                if ws['H523'].value == 'SI':
                    t_infra_instal_1 = constants.SOTANO
                    p_infra_instal_1 = 2
                    g_infra_instal_1 = constants.LEVE
                elif ws['H524'].value == 'SI':
                    t_infra_instal_1 = constants.PRIMER_PISO
                    p_infra_instal_1 = 5
                    g_infra_instal_1 = constants.NULO
                elif ws['H525'].value == 'SI':
                    t_infra_instal_1 = constants.SEGUNDO_PISO
                    p_infra_instal_1 = 2
                    g_infra_instal_1 = constants.LEVE
                elif ws['H526'].value == 'SI':
                    t_infra_instal_1 = constants.OTRO
                    p_infra_instal_1 = 1
                    g_infra_instal_1 = constants.MODERADO
                else:
                    t_infra_instal_1 = False
                    p_infra_instal_1 = 0
                    g_infra_instal_1 = False
                if ws['H527'].value == 'SI':
                    t_infra_instal_2 = constants.DIRECTO
                elif ws['H528'].value == 'SI':
                    t_infra_instal_2 = constants.ESCALERA
                elif ws['H529'].value == 'SI':
                    t_infra_instal_2 = constants.ASCENSOR
                elif ws['H530'].value == 'SI':
                    t_infra_instal_2 = constants.RAMPA
                else:
                    t_infra_instal_2 = False
                if ws['H535'].value == 'SI':
                    infra_instal_6 = constants.SI
                    p_infra_instal_6 = 5
                    g_infra_instal_6 = constants.NULO
                elif ws['H535'].value == 'NO':
                    infra_instal_6 = constants.NO
                    p_infra_instal_6 = 0
                    g_infra_instal_6 = constants.MODERADO
                if ws['H540'].value == 'SI':
                    infra_instal_11 = constants.SI
                    p_infra_instal_11 = 5
                    g_infra_instal_11 = constants.NULO
                elif ws['H540'].value == 'NO':
                    infra_instal_11 = constants.NO
                    p_infra_instal_11 = 0
                    g_infra_instal_11 = constants.MODERADO
                if ws['H543'].value == 'SI':
                    infra_instal_14 = constants.SI
                    p_infra_instal_14 = 5
                    g_infra_instal_14 = constants.NULO
                elif ws['H543'].value == 'NO':
                    infra_instal_14 = constants.NO
                    p_infra_instal_14 = 0
                    g_infra_instal_14 = constants.MODERADO
                if ws['H546'].value == 'SI':
                    infra_instal_17 = constants.SI
                    p_infra_instal_17 = 5
                    g_infra_instal_17 = constants.NULO
                elif ws['H546'].value == 'NO':
                    infra_instal_17 = constants.NO
                    p_infra_instal_17 = 0
                    g_infra_instal_17 = constants.MODERADO
                if ws['H547'].value == 'SI':
                    infra_instal_18 = constants.SI
                    p_infra_instal_18 = 5
                    g_infra_instal_18 = constants.NULO
                elif ws['H547'].value == 'NO':
                    infra_instal_18 = constants.NO
                    p_infra_instal_18 = 0
                    g_infra_instal_18 = constants.MODERADO
                if ws['H548'].value == 'SI':
                    infra_instal_19 = constants.SI
                    p_infra_instal_19 = 5
                    g_infra_instal_19 = constants.NULO
                elif ws['H548'].value == 'NO':
                    infra_instal_19 = constants.NO
                    p_infra_instal_19 = 0
                    g_infra_instal_19 = constants.MODERADO
                if ws['H549'].value == 'SI':
                    infra_instal_20 = constants.SI
                    p_infra_instal_20 = 5
                    g_infra_instal_20 = constants.NULO
                elif ws['H549'].value == 'NO':
                    infra_instal_20 = constants.NO
                    p_infra_instal_20 = 0
                    g_infra_instal_20 = constants.MODERADO
                if ws['H550'].value == 'SI':
                    infra_instal_21 = constants.SI
                    p_infra_instal_21 = 5
                    g_infra_instal_21 = constants.NULO
                elif ws['H550'].value == 'NO':
                    infra_instal_21 = constants.NO
                    p_infra_instal_21 = 0
                    g_infra_instal_21 = constants.MODERADO
                if ws['H551'].value == 'SI':
                    t_infra_instal_21 = constants.PROPIO
                elif ws['H552'].value == 'SI':
                    t_infra_instal_21 = constants.COMPARTIDO
                elif ws['H553'].value == 'SI':
                    t_infra_instal_21 = constants.TERCERIZADO
                else:
                    t_infra_instal_21 = False
                if ws['H554'].value == 'SI':
                    infra_instal_22 = constants.SI
                    p_infra_instal_22 = 5
                    g_infra_instal_22 = constants.NULO
                elif ws['H554'].value == 'NO':
                    infra_instal_22 = constants.NO
                    p_infra_instal_22 = 0
                    g_infra_instal_22 = constants.MODERADO
                if ws['H555'].value == 'SI':
                    infra_instal_23 = constants.SI
                    p_infra_instal_23 = 5
                    g_infra_instal_23 = constants.NULO
                elif ws['H555'].value == 'NO':
                    infra_instal_23 = constants.NO
                    p_infra_instal_23 = 0
                    g_infra_instal_23 = constants.MODERADO
                if ws['H556'].value == 'SI':
                    t_infra_instal_23 = constants.PROPIO
                elif ws['H557'].value == 'SI':
                    t_infra_instal_23 = constants.COMPARTIDO
                elif ws['H558'].value == 'SI':
                    t_infra_instal_23 = constants.TERCERIZADO
                else:
                    t_infra_instal_23 = False
                if ws['H559'].value == 'SI':
                    infra_instal_24 = constants.SI
                elif ws['H559'].value == 'NO':
                    infra_instal_24 = constants.NO
                if ws['H560'].value == 'SI':
                    infra_instal_25 = constants.SI
                elif ws['H560'].value == 'NO':
                    infra_instal_25 = constants.NO
                if ws['H561'].value == 'SI':
                    infra_instal_26 = constants.SI
                    p_infra_instal_26 = 5
                    g_infra_instal_26 = constants.NULO
                elif ws['H561'].value == 'NO':
                    infra_instal_26 = constants.NO
                    p_infra_instal_26 = 0
                    g_infra_instal_26 = constants.MODERADO
                if ws['H562'].value == 'SI':
                    infra_instal_27 = constants.SI
                elif ws['H562'].value == 'NO':
                    infra_instal_27 = constants.NO
                if ws['H563'].value == 'SI':
                    infra_instal_28 = constants.SI
                elif ws['H563'].value == 'NO':
                    infra_instal_28 = constants.NO
                if ws['H566'].value == 'SI':
                    bio_s_2 = constants.SI
                    p_bio_s_2 = 5
                    g_bio_s_2 = constants.NULO
                elif ws['H566'].value == 'NO':
                    bio_s_2 = constants.NO
                    p_bio_s_2 = 0
                    g_bio_s_2 = constants.GRAVE
                if ws['H567'].value == 'SI':
                    bio_s_3 = constants.SI
                elif ws['H567'].value == 'NO':
                    bio_s_3 = constants.NO
                if ws['H568'].value == 'SI':
                    bio_s_4 = constants.SI
                elif ws['H568'].value == 'NO':
                    bio_s_4 = constants.NO
                if ws['H569'].value == 'SI':
                    bio_s_5 = constants.SI
                    p_bio_s_5 = 5
                    g_bio_s_5 = constants.NULO
                elif ws['H569'].value == 'NO':
                    bio_s_5 = constants.NO
                    p_bio_s_5 = 0
                    g_bio_s_5 = constants.GRAVE

                ficha_id.write({
                    'rg_1': rg_1,
                    'obs_rg_1': ws['J46'].value,
                    'p_rg_1': p_rg_1,
                    'g_rg_1': g_rg_1,
                    'rg_2': rg_2,
                    'obs_rg_2': ws['J47'].value,
                    'p_rg_2': p_rg_2,
                    'g_rg_2': g_rg_2,
                    't_rg_3': t_rg_3,
                    'obs_rg_3': ws['J48'].value,
                    'p_rg_3': p_rg_3,
                    'rg_4': rg_4,
                    'obs_rg_4': ws['J51'].value,
                    'p_rg_4': p_rg_4,
                    'g_rg_4': g_rg_4,
                    'rg_5': rg_5,
                    'obs_rg_5': ws['J52'].value,
                    'p_rg_5': p_rg_5,
                    'g_rg_5': g_rg_5,
                    'rg_6': rg_6,
                    'obs_rg_6': ws['J53'].value,
                    'p_rg_6': p_rg_6,
                    'g_rg_6': g_rg_6,
                    'rg_8': rg_8,
                    'obs_rg_8': ws['J54'].value,
                    'p_rg_8': p_rg_8,
                    'g_rg_8': g_rg_8,
                    'rg_9': rg_9,
                    'obs_rg_9': ws['J55'].value,
                    'p_rg_9': p_rg_9,
                    'g_rg_9': g_rg_9,
                    'rg_10': rg_10,
                    'obs_rg_10': ws['J56'].value,
                    'p_rg_10': p_rg_10,
                    'g_rg_10': g_rg_10,
                    'rg_11': rg_11,
                    'obs_rg_11': ws['J57'].value,
                    'p_rg_11': p_rg_11,
                    'g_rg_11': g_rg_11,
                    'rg_12': rg_12,
                    'obs_rg_12': ws['J58'].value,
                    'p_rg_12': p_rg_12,
                    'g_rg_12': g_rg_12,
                    'rg_13': rg_13,
                    'obs_rg_13': ws['J59'].value,
                    'p_rg_13': p_rg_13,
                    'g_rg_13': g_rg_13,
                    'rg_14': rg_14,
                    'obs_rg_14': ws['J60'].value,
                    'p_rg_14': p_rg_14,
                    'g_rg_14': g_rg_14,
                    'rg_15': rg_15,
                    'obs_rg_15': ws['J61'].value,
                    'p_rg_15': p_rg_15,
                    'g_rg_15': g_rg_15,
                    'rg_16': rg_16,
                    'obs_rg_16': ws['J62'].value,
                    'p_rg_16': p_rg_16,
                    'g_rg_16': g_rg_16,
                    'rg_17': rg_17,
                    'obs_rg_17': ws['J63'].value,
                    'esp_rg_17': ws['I63'].value,
                    'p_rg_17': p_rg_17,
                    'g_rg_17': g_rg_17,
                    'rg_18': rg_18,
                    'obs_rg_18': ws['J64'].value,
                    'p_rg_18': p_rg_18,
                    'g_rg_18': g_rg_18,
                    'rg_19': rg_19,
                    'obs_rg_19': ws['J66'].value,
                    'p_rg_19': p_rg_19,
                    'g_rg_19': g_rg_19,
                    't_rg_20': t_rg_20,
                    'obs_rg_20': ws['J67'].value,
                    'p_rg_20': p_rg_20,
                    'g_rg_20': g_rg_20,
                    'rg_21': rg_21,
                    'obs_rg_21': ws['J71'].value,
                    'p_rg_21': p_rg_21,
                    'g_rg_21': g_rg_21,
                    't_rg_22': t_rg_22,
                    'obs_rg_22': ws['J72'].value,
                    'rg_23': rg_23,
                    'obs_rg_23': ws['J78'].value,
                    'p_rg_23': p_rg_23,
                    'g_rg_23': g_rg_23,
                    'rg_24': rg_24,
                    'obs_rg_24': ws['J79'].value,
                    'p_rg_24': p_rg_24,
                    'g_rg_24': g_rg_24,
                    'rg_25': rg_25,
                    'obs_rg_25': ws['J80'].value,
                    'p_rg_25': p_rg_25,
                    'g_rg_25': g_rg_25,
                    'rg_26': rg_26,
                    'obs_rg_26': ws['J81'].value,
                    'p_rg_26': p_rg_26,
                    'g_rg_26': g_rg_26,
                    'rg_28': rg_28,
                    'obs_rg_28': ws['J82'].value,
                    'p_rg_28': p_rg_28,
                    'g_rg_28': g_rg_28,
                    'rg_29': rg_29,
                    'obs_rg_29': ws['J83'].value,
                    'p_rg_29': p_rg_29,
                    'g_rg_29': g_rg_29,
                    'rg_30': rg_30,
                    'obs_rg_30': ws['J84'].value,
                    'p_rg_30': p_rg_30,
                    'g_rg_30': g_rg_30,
                    'rh_1': rh_1,
                    'esp_rh_1': ws['I88'].value,
                    'obs_rh_1': ws['J88'].value,
                    'p_rh_1': p_rh_1,
                    'g_rh_1': g_rh_1,
                    'rh_2': rh_2,
                    'esp_rh_2': ws['I89'].value,
                    'obs_rh_2': ws['J89'].value,
                    'p_rh_2': p_rh_2,
                    'g_rh_2': g_rh_2,
                    'rh_3': rh_3,
                    'esp_rh_3': ws['I90'].value,
                    'obs_rh_3': ws['J90'].value,
                    'p_rh_3': p_rh_3,
                    'g_rh_3': g_rh_3,
                    'rh_4': rh_4,
                    'esp_rh_4': ws['I91'].value,
                    'obs_rh_4': ws['J91'].value,
                    'p_rh_4': p_rh_4,
                    'g_rh_4': g_rh_4,
                    'rh_5': rh_5,
                    'esp_rh_5': ws['I92'].value,
                    'obs_rh_5': ws['J92'].value,
                    'p_rh_5': p_rh_5,
                    'g_rh_5': g_rh_5,
                    'rh_6': rh_6,
                    'esp_rh_6': ws['I93'].value,
                    'obs_rh_6': ws['J93'].value,
                    'p_rh_6': p_rh_6,
                    'g_rh_6': g_rh_6,
                    'rh_7': rh_7,
                    'esp_rh_7': ws['I94'].value,
                    'obs_rh_7': ws['J94'].value,
                    'p_rh_7': p_rh_7,
                    'g_rh_7': g_rh_7,
                    'rh_8': rh_8,
                    'esp_rh_8': ws['I95'].value,
                    'obs_rh_8': ws['J95'].value,
                    'p_rh_8': p_rh_8,
                    'g_rh_8': g_rh_8,
                    'rh_9': rh_9,
                    'esp_rh_9': ws['I96'].value,
                    'obs_rh_9': ws['J96'].value,
                    'p_rh_9': p_rh_9,
                    'g_rh_9': g_rh_9,
                    'rh_10': rh_10,
                    'obs_rh_10': ws['J97'].value,
                    'p_rh_10': p_rh_10,
                    'g_rh_10': g_rh_10,
                    'obs_pros_1': ws['J102'].value,
                    'esp_pros_1': ws['I102'].value,
                    'pros_2': pros_2,
                    'obs_pros_2': ws['J103'].value,
                    'p_pros_2': p_pros_2,
                    'g_pros_2': g_pros_2,
                    'pros_3': pros_3,
                    'obs_pros_3': ws['J104'].value,
                    'p_pros_3': p_pros_3,
                    'g_pros_3': g_pros_3,
                    't_pros_4': t_pros_4,
                    'obs_pros_4': ws['J105'].value,
                    'pros_5': pros_5,
                    'obs_pros_5': ws['J108'].value,
                    'p_pros_5': p_pros_5,
                    'g_pros_5': g_pros_5,
                    'pros_6': pros_6,
                    'obs_pros_6': ws['J109'].value,
                    'p_pros_6': p_pros_6,
                    'g_pros_6': g_pros_6,
                    't_pros_7': t_pros_7,
                    'obs_pros_7': ws['J110'].value,
                    'p_pros_7': p_pros_7,
                    'g_pros_7': g_pros_7,
                    't_pros_8': t_pros_8,
                    'obs_pros_8': ws['J112'].value,
                    'p_pros_8': p_pros_8,
                    'g_pros_8': g_pros_8,
                    'pros_9': pros_9,
                    'obs_pros_9': ws['J115'].value,
                    'p_pros_9': p_pros_9,
                    'g_pros_9': g_pros_9,
                    'pros_10': pros_10,
                    'obs_pros_10': ws['J116'].value,
                    'p_pros_10': p_pros_10,
                    'g_pros_10': g_pros_10,
                    'pros_11': pros_11,
                    'obs_pros_11': ws['J117'].value,
                    'p_pros_11': p_pros_11,
                    'g_pros_11': g_pros_11,
                    't_pros_12': t_pros_12,
                    'obs_pros_12': ws['J118'].value,
                    'p_pros_12': p_pros_12,
                    'g_pros_12': g_pros_12,
                    'pros_13': pros_13,
                    'obs_pros_13': ws['J120'].value,
                    'p_pros_13': p_pros_13,
                    'g_pros_13': g_pros_13,
                    'pros_14': pros_14,
                    'obs_pros_14': ws['J121'].value,
                    'p_pros_14': p_pros_14,
                    'g_pros_14': g_pros_14,
                    'pros_15': pros_15,
                    'obs_pros_15': ws['J122'].value,
                    'p_pros_15': p_pros_15,
                    'g_pros_15': g_pros_15,
                    'pros_16': pros_16,
                    'obs_pros_16': ws['J123'].value,
                    'p_pros_16': p_pros_16,
                    'g_pros_16': g_pros_16,
                    'pros_17': pros_17,
                    't_pros_17': t_pros_17,
                    'obs_pros_17': ws['J124'].value,
                    'p_pros_17': p_pros_17,
                    'g_pros_17': g_pros_17,
                    'pros_18': pros_18,
                    'obs_pros_18': ws['J128'].value,
                    'p_pros_18': p_pros_18,
                    'g_pros_18': g_pros_18,
                    't_pros_19': t_pros_19,
                    'obs_pros_19': ws['J129'].value,
                    'esp_pros_20': ws['I131'].value,
                    'obs_pros_20': ws['J131'].value,
                    'pros_21': pros_21,
                    'obs_pros_21': ws['J133'].value,
                    'p_pros_21': p_pros_21,
                    'g_pros_21': g_pros_21,
                    'pros_22': pros_22,
                    'obs_pros_22': ws['J134'].value,
                    'p_pros_22': p_pros_22,
                    'g_pros_22': g_pros_22,
                    'obs_pros_23': ws['J135'].value,
                    't_pros_23': t_pros_23,
                    'obs_pros_24': ws['J137'].value,
                    't_pros_24': t_pros_24,
                    'pros_33': pros_33,
                    'obs_pros_33': ws['J140'].value,
                    'pros_25': pros_25,
                    'obs_pros_25': ws['J142'].value,
                    'pros_26': pros_26,
                    'obs_pros_26': ws['J143'].value,
                    'pros_27': pros_27,
                    'obs_pros_27': ws['J144'].value,
                    'pros_28': pros_28,
                    'obs_pros_28': ws['J145'].value,
                    'pros_29': pros_29,
                    'obs_pros_29': ws['J146'].value,
                    'pros_30': pros_30,
                    'obs_pros_30': ws['J147'].value,
                    'pros_31': pros_31,
                    'obs_pros_31': ws['J148'].value,
                    'pros_32': pros_32,
                    'obs_pros_32': ws['J149'].value,
                    'pros_34': pros_34,
                    'obs_pros_34': ws['J151'].value,
                    'p_pros_34': p_pros_34,
                    'g_pros_34': g_pros_34,
                    't_pros_34': t_pros_34,
                    'pros_35': pros_35,
                    'obs_pros_35': ws['J154'].value,
                    'p_pros_35': p_pros_35,
                    'g_pros_35': g_pros_35,
                    'pros_36': pros_36,
                    'obs_pros_36': ws['J155'].value,
                    'p_pros_36': p_pros_36,
                    'g_pros_36': g_pros_36,
                    'pros_37': pros_37,
                    'obs_pros_37': ws['J157'].value,
                    'p_pros_37': p_pros_37,
                    'g_pros_37': g_pros_37,
                    'pros_38': pros_38,
                    'obs_pros_38': ws['J158'].value,
                    'p_pros_38': p_pros_38,
                    'g_pros_38': g_pros_38,
                    't_pros_38': t_pros_38,
                    'pros_39': pros_39,
                    'obs_pros_39': ws['J161'].value,
                    'p_pros_39': p_pros_39,
                    'g_pros_39': g_pros_39,
                    't_pros_39': t_pros_39,
                    'pros_40': pros_40,
                    'obs_pros_40': ws['J164'].value,
                    'p_pros_40': p_pros_40,
                    'g_pros_40': g_pros_40,
                    't_pros_40': t_pros_40,
                    'pros_41': pros_41,
                    'obs_pros_41': ws['J167'].value,
                    'p_pros_41': p_pros_41,
                    'g_pros_41': g_pros_41,
                    't_pros_41': t_pros_41,
                    'pros_42': pros_42,
                    'obs_pros_42': ws['J170'].value,
                    'p_pros_42': p_pros_42,
                    'g_pros_42': g_pros_42,
                    't_pros_42': t_pros_42,
                    'pros_43': pros_43,
                    'obs_pros_43': ws['J173'].value,
                    'p_pros_43': p_pros_43,
                    'g_pros_43': g_pros_43,
                    't_pros_43': t_pros_43,
                    'pros_44': pros_44,
                    'obs_pros_44': ws['J176'].value,
                    'p_pros_44': p_pros_44,
                    'g_pros_44': g_pros_44,
                    't_pros_44': t_pros_44,
                    'pros_45': pros_45,
                    'obs_pros_45': ws['J179'].value,
                    'esp_pros_45': ws['I179'].value,
                    'obs_pros_47': ws['J180'].value,
                    't_pros_47': t_pros_47,
                    'esp_pros_48': ws['I183'].value,
                    'obs_pros_48': ws['J183'].value,
                    'pros_49': pros_49,
                    'obs_pros_49': ws['J184'].value,
                    'p_pros_49': p_pros_49,
                    'g_pros_49': g_pros_49,
                    'pros_50': pros_50,
                    'obs_pros_50': ws['J185'].value,
                    'p_pros_50': p_pros_50,
                    'g_pros_50': g_pros_50,
                    't_pros_50': t_pros_50,
                    'pros_51': pros_51,
                    'obs_pros_51': ws['J189'].value,
                    'p_pros_51': p_pros_51,
                    'g_pros_51': g_pros_51,
                    'pros_52': pros_52,
                    'obs_pros_52': ws['J190'].value,
                    'p_pros_52': p_pros_52,
                    'g_pros_52': g_pros_52,
                    'pros_53': pros_53,
                    'obs_pros_53': ws['J191'].value,
                    'p_pros_53': p_pros_53,
                    'g_pros_53': g_pros_53,
                    'r_pros_53': r_pros_53,
                    'r_obs_53': ws['J192'].value,
                    'p_r_pros_53': p_r_pros_53,
                    'g_r_pros_53': g_r_pros_53,
                    'pros_54': pros_54,
                    'obs_pros_54': ws['J193'].value,
                    'p_pros_54': p_pros_54,
                    'g_pros_54': g_pros_54,
                    'pros_55': pros_55,
                    'obs_pros_55': ws['J194'].value,
                    'p_pros_55': p_pros_55,
                    'g_pros_55': g_pros_55,
                    'obs_pros_56': ws['J196'].value,
                    't_pros_56': t_pros_56,
                    'pros_57': pros_57,
                    'obs_pros_57': ws['J199'].value,
                    'p_pros_57': p_pros_57,
                    'g_pros_57': g_pros_57,
                    'pros_58': pros_58,
                    'obs_pros_58': ws['J200'].value,
                    'p_pros_58': p_pros_58,
                    'g_pros_58': g_pros_58,
                    'pros_59': pros_59,
                    'obs_pros_59': ws['J201'].value,
                    'p_pros_59': p_pros_59,
                    'g_pros_59': g_pros_59,
                    'pros_60': pros_60,
                    'obs_pros_60': ws['J202'].value,
                    'p_pros_60': p_pros_60,
                    'g_pros_60': g_pros_60,
                    'pros_61': pros_61,
                    'obs_pros_61': ws['J203'].value,
                    'p_pros_61': p_pros_61,
                    'g_pros_61': g_pros_61,
                    'pros_62': pros_62,
                    'obs_pros_62': ws['J204'].value,
                    'p_pros_62': p_pros_62,
                    'g_pros_62': g_pros_62,
                    'pros_63': pros_63,
                    'obs_pros_63': ws['J205'].value,
                    'p_pros_63': p_pros_63,
                    'g_pros_63': g_pros_63,
                    'pros_64': pros_64,
                    'obs_pros_64': ws['J206'].value,
                    'p_pros_64': p_pros_64,
                    'g_pros_64': g_pros_64,
                    'pros_65': pros_65,
                    'obs_pros_65': ws['J207'].value,
                    'p_pros_65': p_pros_65,
                    'g_pros_65': g_pros_65,
                    'pros_66': pros_66,
                    'esp_pros_66': ws['I208'].value,
                    'obs_pros_66': ws['J208'].value,
                    'p_pros_66': p_pros_66,
                    'g_pros_66': g_pros_66,
                    'pros_68': pros_68,
                    't_pros_68': t_pros_68,
                    'obs_pros_68': ws['J209'].value,
                    'pros_69': pros_69,
                    'obs_pros_69': ws['J215'].value,
                    'p_pros_69': p_pros_69,
                    'g_pros_69': g_pros_69,
                    'pros_70': pros_70,
                    'obs_pros_70': ws['J216'].value,
                    'p_pros_70': p_pros_70,
                    'g_pros_70': g_pros_70,
                    'pros_73': pros_73,
                    'obs_pros_73': ws['J218'].value,
                    'p_pros_73': p_pros_73,
                    'g_pros_73': g_pros_73,
                    'pros_74': pros_74,
                    'obs_pros_74': ws['J219'].value,
                    'p_pros_74': p_pros_74,
                    'g_pros_74': g_pros_74,
                    'pros_75': pros_75,
                    'obs_pros_75': ws['J220'].value,
                    'p_pros_75': p_pros_75,
                    'g_pros_75': g_pros_75,
                    'pros_76': pros_76,
                    'obs_pros_76': ws['J221'].value,
                    'p_pros_76': p_pros_76,
                    'g_pros_76': g_pros_76,
                    'pros_77': pros_77,
                    'obs_pros_77': ws['J222'].value,
                    'p_pros_77': p_pros_77,
                    'g_pros_77': g_pros_77,
                    't_pros_77': t_pros_77,
                    'pros_78': pros_78,
                    'obs_pros_78': ws['J227'].value,
                    'p_pros_78': p_pros_78,
                    'g_pros_78': g_pros_78,
                    'pros_79': pros_79,
                    'obs_pros_79': ws['J229'].value,
                    'p_pros_79': p_pros_79,
                    'g_pros_79': g_pros_79,
                    'pros_80': pros_80,
                    'obs_pros_80': ws['J230'].value,
                    'p_pros_80': p_pros_80,
                    'g_pros_80': g_pros_80,
                    'pros_81': pros_81,
                    'obs_pros_81': ws['J233'].value,
                    'pros_82': pros_82,
                    'obs_pros_82': ws['J234'].value,
                    'pros_83': pros_83,
                    'obs_pros_83': ws['J235'].value,
                    'pros_84': pros_84,
                    'obs_pros_84': ws['J236'].value,
                    'pros_85': pros_85,
                    'obs_pros_85': ws['J238'].value,
                    'p_pros_85': p_pros_85,
                    'g_pros_85': g_pros_85,
                    'pros_87': pros_87,
                    'obs_pros_87': ws['J240'].value,
                    'pros_88': pros_88,
                    'obs_pros_88': ws['J242'].value,
                    'pros_89': pros_89,
                    'obs_pros_89': ws['J243'].value,
                    'pros_90': pros_90,
                    'obs_pros_90': ws['J244'].value,
                    'pros_91': pros_91,
                    'obs_pros_91': ws['J245'].value,
                    'pros_92': pros_92,
                    'obs_pros_92': ws['J246'].value,
                    'pros_93': pros_93,
                    'obs_pros_93': ws['J248'].value,
                    'p_pros_93': p_pros_93,
                    'g_pros_93': g_pros_93,
                    't_pros_93': t_pros_93,
                    'pros_94': pros_94,
                    'obs_pros_94': ws['J258'].value,
                    'p_pros_94': p_pros_94,
                    'g_pros_94': g_pros_94,
                    'pros_95': pros_95,
                    'obs_pros_95': ws['J259'].value,
                    'p_pros_95': p_pros_95,
                    'g_pros_95': g_pros_95,
                    'pros_96': pros_96,
                    'obs_pros_96': ws['J260'].value,
                    'pros_97': pros_97,
                    'obs_pros_97': ws['J261'].value,
                    'pros_98': pros_98,
                    'obs_pros_98': ws['J262'].value,
                    'p_pros_98': p_pros_98,
                    'g_pros_98': g_pros_98,
                    'pros_99': pros_99,
                    'obs_pros_99': ws['J252'].value,
                    'p_pros_99': p_pros_99,
                    'g_pros_99': g_pros_99,
                    'pros_100': pros_100,
                    'obs_pros_100': ws['J253'].value,
                    'p_pros_100': p_pros_100,
                    'g_pros_100': g_pros_100,
                    'pros_101': pros_101,
                    'obs_pros_101': ws['J254'].value,
                    'p_pros_101': p_pros_101,
                    'g_pros_101': g_pros_101,
                    'pros_102': pros_102,
                    'obs_pros_102': ws['J255'].value,
                    'p_pros_102': p_pros_102,
                    'g_pros_102': g_pros_102,
                    'pros_103': pros_103,
                    'obs_pros_103': ws['J256'].value,
                    'p_pros_103': p_pros_103,
                    'g_pros_103': g_pros_103,
                    'cali_1': cali_1,
                    'obs_cali_1': ws['J266'].value,
                    'cali_2': cali_2,
                    'obs_cali_2': ws['J267'].value,
                    't_cali_2': t_cali_2,
                    'cali_3': cali_3,
                    'obs_cali_3': ws['J268'].value,
                    'cali_4': cali_4,
                    'obs_cali_4': ws['J270'].value,
                    'cali_5': cali_5,
                    'obs_cali_5': ws['J271'].value,
                    'cali_6': cali_6,
                    'obs_cali_6': ws['J272'].value,
                    'cali_7': cali_7,
                    'obs_cali_7': ws['J273'].value,
                    'cali_8': cali_8,
                    'obs_cali_8': ws['J274'].value,
                    'esp_alm_1': ws['I284'].value,
                    'obs_alm_1': ws['J284'].value,
                    'esp_alm_2': ws['I285'].value,
                    'obs_alm_2': ws['J285'].value,
                    'esp_alm_3': ws['I286'].value,
                    'obs_alm_3': ws['J286'].value,
                    'esp_alm_4': ws['I287'].value,
                    'obs_alm_4': ws['J287'].value,
                    'esp_alm_5': ws['I288'].value,
                    'obs_alm_5': ws['J288'].value,
                    'alm_6': alm_6,
                    'obs_alm_6': ws['J289'].value,
                    'p_alm_6': p_alm_6,
                    'g_alm_6': g_alm_6,
                    'alm_7': alm_7,
                    'obs_alm_7': ws['J290'].value,
                    'p_alm_7': p_alm_7,
                    'g_alm_7': g_alm_7,
                    'alm_8': alm_8,
                    'obs_alm_8': ws['J291'].value,
                    'p_alm_8': p_alm_8,
                    'g_alm_8': g_alm_8,
                    'esp_alm_9': ws['I293'].value,
                    'obs_alm_9': ws['J293'].value,
                    'g_alm_9': constants.NULO,
                    'esp_alm_10 ': ws['I294'].value,
                    'obs_alm_10 ': ws['J294'].value,
                    'g_alm_10 ': constants.NULO,
                    'esp_alm_11 ': ws['I295'].value,
                    'obs_alm_11 ': ws['J295'].value,
                    'esp_alm_12 ': ws['I296'].value,
                    'obs_alm_12 ': ws['J296'].value,
                    'esp_alm_13': ws['I297'].value,
                    'obs_alm_13 ': ws['J297'].value,
                    'esp_alm_14': ws['I298'].value,
                    'obs_alm_14 ': ws['J298'].value,
                    'sis_inform_1': sis_inform_1,
                    'obs_sis_inform_1': ws['J301'].value,
                    'esp_sis_inform_2': ws['I302'].value,
                    'obs_sis_inform_2': ws['J302'].value,
                    'esp_sis_inform_3': ws['I303'].value,
                    'obs_sis_inform_3': ws['J303'].value,
                    't_sis_inform_4': t_sis_inform_4,
                    'obs_sis_inform_4': ws['J304'].value,
                    'sis_inform_5': sis_inform_5,
                    'obs_sis_inform_5': ws['J306'].value,
                    'clasf_bio_2': clasf_bio_2,
                    'obs_clasf_bio_2': ws['J310'].value,
                    'p_clasf_bio_2': p_clasf_bio_2,
                    'g_clasf_bio_2': g_clasf_bio_2,
                    'clasf_bio_3': clasf_bio_3,
                    'obs_clasf_bio_3': ws['J311'].value,
                    'p_clasf_bio_3': p_clasf_bio_3,
                    'g_clasf_bio_3': g_clasf_bio_3,
                    'esp_clasf_bio_4': ws['I312'].value,
                    'obs_clasf_bio_4 ': ws['J312'].value,
                    't_clasf_bio_5': t_clasf_bio_5,
                    'obs_clasf_bio_5 ': ws['J313'].value,
                    'clasf_bio_6': clasf_bio_6,
                    'obs_clasf_bio_6': ws['J316'].value,
                    'p_clasf_bio_6': p_clasf_bio_6,
                    'g_clasf_bio_6': g_clasf_bio_6,
                    'clasf_bio_7': clasf_bio_7,
                    'obs_clasf_bio_7': ws['J317'].value,
                    'p_clasf_bio_7': p_clasf_bio_7,
                    'g_clasf_bio_7': g_clasf_bio_7,
                    'clasf_bio_8': clasf_bio_8,
                    'obs_clasf_bio_8': ws['J318'].value,
                    'p_clasf_bio_8': p_clasf_bio_8,
                    'g_clasf_bio_8': g_clasf_bio_8,
                    't_clasf_bio_9': t_clasf_bio_9,
                    'obs_clasf_bio_9': ws['J319'].value,
                    'p_clasf_bio_9': p_clasf_bio_9,
                    'g_clasf_bio_9': g_clasf_bio_9,
                    'clasf_bio_14': clasf_bio_14,
                    'obs_clasf_bio_14': ws['J325'].value,
                    'p_clasf_bio_14': p_clasf_bio_14,
                    'g_clasf_bio_14': g_clasf_bio_14,
                    'clasf_bio_15': clasf_bio_15,
                    'obs_clasf_bio_15': ws['J326'].value,
                    'p_clasf_bio_15': p_clasf_bio_15,
                    'g_clasf_bio_15': g_clasf_bio_15,
                    'esp_clasf_bio_16': ws['I327'].value,
                    'obs_clasf_bio_16 ': ws['J327'].value,
                    'clasf_bio_18': clasf_bio_18,
                    'obs_clasf_bio_18': ws['J328'].value,
                    'p_clasf_bio_18': p_clasf_bio_18,
                    'g_clasf_bio_18': g_clasf_bio_18,
                    'clasf_bio_19': clasf_bio_19,
                    'obs_clasf_bio_19': ws['J329'].value,
                    'p_clasf_bio_19': p_clasf_bio_19,
                    'g_clasf_bio_19': g_clasf_bio_19,
                    'equip_1': equip_1,
                    'obs_equip_1': ws['J332'].value,
                    'p_equip_1': p_equip_1,
                    'g_equip_1': g_equip_1,
                    'equip_2': equip_2,
                    'obs_equip_2': ws['J333'].value,
                    'p_equip_2': p_equip_2,
                    'g_equip_2': g_equip_2,
                    'equip_3': equip_3,
                    'obs_equip_3': ws['J334'].value,
                    'p_equip_3': p_equip_3,
                    'g_equip_3': g_equip_3,
                    'equip_4': equip_4,
                    'obs_equip_4': ws['J335'].value,
                    'p_equip_4': p_equip_4,
                    'g_equip_4': g_equip_4,
                    'equip_5': equip_5,
                    'obs_equip_5': ws['J336'].value,
                    'p_equip_5': p_equip_5,
                    'g_equip_5': g_equip_5,
                    'equip_6': equip_6,
                    'obs_equip_6': ws['J337'].value,
                    'p_equip_6': p_equip_6,
                    'g_equip_6': g_equip_6,
                    'equip_7': equip_7,
                    'obs_equip_7': ws['J338'].value,
                    'p_equip_7': p_equip_7,
                    'g_equip_7': g_equip_7,
                    'equip_8': equip_8,
                    'obs_equip_8': ws['J339'].value,
                    'p_equip_8': p_equip_8,
                    'g_equip_8': g_equip_8,
                    'equip_9': equip_9,
                    'obs_equip_9': ws['J340'].value,
                    'p_equip_9': p_equip_9,
                    'g_equip_9': g_equip_9,
                    'equip_10': equip_10,
                    'obs_equip_10': ws['J341'].value,
                    'p_equip_10': p_equip_10,
                    'g_equip_10': g_equip_10,
                    'equip_11': equip_11,
                    'obs_equip_11': ws['J342'].value,
                    'p_equip_11': p_equip_11,
                    'g_equip_11': g_equip_11,
                    'rel_equip_cant_c_1': ws['E347'].value,
                    'rel_equip_cant_c_b_1': ws['F348'].value,
                    'rel_equip_cant_c_r_1': ws['F349'].value,
                    'rel_equip_cant_c_m_1': ws['F350'].value,
                    'rel_c_equip_1': rel_c_equip_1,
                    'rel_equip_cant_p_1': ws['G347'].value,
                    'rel_equip_cant_p_b_1': ws['H348'].value,
                    'rel_equip_cant_p_r_1': ws['H349'].value,
                    'rel_equip_cant_p_m_1': ws['H350'].value,
                    'rel_p_equip_1': rel_p_equip_1,
                    'obs_rel_equip_1': ws['I347'].value,
                    'rel_equip_cant_c_2': ws['E352'].value,
                    'rel_equip_cant_c_b_2': ws['F353'].value,
                    'rel_equip_cant_c_r_2': ws['F354'].value,
                    'rel_equip_cant_c_m_2': ws['F355'].value,
                    'rel_c_equip_2': rel_c_equip_2,
                    'rel_equip_cant_p_2': ws['G352'].value,
                    'rel_equip_cant_p_b_2': ws['H353'].value,
                    'rel_equip_cant_p_r_2': ws['H354'].value,
                    'rel_equip_cant_p_m_2': ws['H355'].value,
                    'rel_p_equip_2': rel_p_equip_2,
                    'obs_rel_equip_2': ws['I352'].value,
                    'rel_equip_cant_c_3': ws['E357'].value,
                    'rel_equip_cant_c_b_3': ws['F358'].value,
                    'rel_equip_cant_c_r_3': ws['F359'].value,
                    'rel_equip_cant_c_m_3': ws['F360'].value,
                    'rel_c_equip_3': rel_c_equip_3,
                    'rel_equip_cant_p_3': ws['G357'].value,
                    'rel_equip_cant_p_b_3': ws['H358'].value,
                    'rel_equip_cant_p_r_3': ws['H359'].value,
                    'rel_equip_cant_p_m_3': ws['H360'].value,
                    'rel_p_equip_3': rel_p_equip_3,
                    'obs_rel_equip_3': ws['I357'].value,
                    'rel_equip_cant_c_4': ws['E362'].value,
                    'rel_equip_cant_c_b_4': ws['F363'].value,
                    'rel_equip_cant_c_r_4': ws['F364'].value,
                    'rel_equip_cant_c_m_4': ws['F365'].value,
                    'rel_c_equip_4': rel_c_equip_4,
                    'rel_equip_cant_p_4': ws['G362'].value,
                    'rel_equip_cant_p_b_4': ws['H363'].value,
                    'rel_equip_cant_p_r_4': ws['H364'].value,
                    'rel_equip_cant_p_m_4': ws['H365'].value,
                    'rel_p_equip_4': rel_p_equip_4,
                    'obs_rel_equip_4': ws['I362'].value,
                    'rel_equip_cant_c_5': ws['E367'].value,
                    'rel_equip_cant_c_b_5': ws['F368'].value,
                    'rel_equip_cant_c_r_5': ws['F369'].value,
                    'rel_equip_cant_c_m_5': ws['F370'].value,
                    'rel_c_equip_5': rel_c_equip_5,
                    'rel_equip_cant_p_5': ws['G367'].value,
                    'rel_equip_cant_p_b_5': ws['H368'].value,
                    'rel_equip_cant_p_r_5': ws['H369'].value,
                    'rel_equip_cant_p_m_5': ws['H370'].value,
                    'rel_p_equip_5': rel_p_equip_5,
                    'obs_rel_equip_5': ws['I367'].value,
                    'rel_equip_cant_c_6': ws['E372'].value,
                    'rel_equip_cant_c_b_6': ws['F373'].value,
                    'rel_equip_cant_c_r_6': ws['F374'].value,
                    'rel_equip_cant_c_m_6': ws['F375'].value,
                    'rel_c_equip_6': rel_c_equip_6,
                    'rel_equip_cant_p_6': ws['G372'].value,
                    'rel_equip_cant_p_b_6': ws['H373'].value,
                    'rel_equip_cant_p_r_6': ws['H374'].value,
                    'rel_equip_cant_p_m_6': ws['H375'].value,
                    'rel_p_equip_6': rel_p_equip_6,
                    'obs_rel_equip_6': ws['I372'].value,
                    'rel_equip_cant_c_7': ws['E377'].value,
                    'rel_equip_cant_c_b_7': ws['F378'].value,
                    'rel_equip_cant_c_r_7': ws['F379'].value,
                    'rel_equip_cant_c_m_7': ws['F380'].value,
                    'rel_c_equip_7': rel_c_equip_7,
                    'rel_equip_cant_p_7': ws['G377'].value,
                    'rel_equip_cant_p_b_7': ws['H378'].value,
                    'rel_equip_cant_p_r_7': ws['H379'].value,
                    'rel_equip_cant_p_m_7': ws['H380'].value,
                    'rel_p_equip_7': rel_p_equip_7,
                    'obs_rel_equip_7': ws['I377'].value,
                    'rel_equip_cant_c_8': ws['E382'].value,
                    'rel_equip_cant_c_b_8': ws['F383'].value,
                    'rel_equip_cant_c_r_8': ws['F384'].value,
                    'rel_equip_cant_c_m_8': ws['F385'].value,
                    'rel_c_equip_8': rel_c_equip_8,
                    'rel_equip_cant_p_8': ws['G382'].value,
                    'rel_equip_cant_p_b_8': ws['H383'].value,
                    'rel_equip_cant_p_r_8': ws['H384'].value,
                    'rel_equip_cant_p_m_8': ws['H385'].value,
                    'rel_p_equip_8': rel_p_equip_8,
                    'obs_rel_equip_8': ws['I382'].value,
                    'rel_equip_cant_c_9': ws['E387'].value,
                    'rel_equip_cant_c_b_9': ws['F388'].value,
                    'rel_equip_cant_c_r_9': ws['F389'].value,
                    'rel_equip_cant_c_m_9': ws['F390'].value,
                    'rel_c_equip_9': rel_c_equip_9,
                    'rel_equip_cant_p_9': ws['G387'].value,
                    'rel_equip_cant_p_b_9': ws['H388'].value,
                    'rel_equip_cant_p_r_9': ws['H389'].value,
                    'rel_equip_cant_p_m_9': ws['H390'].value,
                    'rel_p_equip_9': rel_p_equip_9,
                    'obs_rel_equip_9': ws['I387'].value,
                    'rel_equip_cant_c_10': ws['E392'].value,
                    'rel_equip_cant_c_b_10': ws['F393'].value,
                    'rel_equip_cant_c_r_10': ws['F394'].value,
                    'rel_equip_cant_c_m_10': ws['F395'].value,
                    'rel_c_equip_10': rel_c_equip_10,
                    'rel_equip_cant_p_10': ws['G392'].value,
                    'rel_equip_cant_p_b_10': ws['H393'].value,
                    'rel_equip_cant_p_r_10': ws['H394'].value,
                    'rel_equip_cant_p_m_10': ws['H395'].value,
                    'rel_p_equip_10': rel_p_equip_10,
                    'obs_rel_equip_10': ws['I392'].value,
                    'rel_equip_cant_c_11': ws['E397'].value,
                    'rel_equip_cant_c_b_11': ws['F398'].value,
                    'rel_equip_cant_c_r_11': ws['F399'].value,
                    'rel_equip_cant_c_m_11': ws['F400'].value,
                    'rel_c_equip_11': rel_c_equip_11,
                    'rel_equip_cant_p_11': ws['G397'].value,
                    'rel_equip_cant_p_b_11': ws['H398'].value,
                    'rel_equip_cant_p_r_11': ws['H399'].value,
                    'rel_equip_cant_p_m_11': ws['H400'].value,
                    'rel_p_equip_11': rel_p_equip_11,
                    'obs_rel_equip_11': ws['I397'].value,
                    'rel_equip_cant_c_12': ws['E402'].value,
                    'rel_equip_cant_c_b_12': ws['F403'].value,
                    'rel_equip_cant_c_r_12': ws['F404'].value,
                    'rel_equip_cant_c_m_12': ws['F405'].value,
                    'rel_c_equip_12': rel_c_equip_12,
                    'rel_equip_cant_p_12': ws['G402'].value,
                    'rel_equip_cant_p_b_12': ws['H403'].value,
                    'rel_equip_cant_p_r_12': ws['H404'].value,
                    'rel_equip_cant_p_m_12': ws['H405'].value,
                    'rel_p_equip_12': rel_p_equip_12,
                    'obs_rel_equip_12': ws['I402'].value,
                    'rel_equip_cant_c_13': ws['E407'].value,
                    'rel_equip_cant_c_b_13': ws['F408'].value,
                    'rel_equip_cant_c_r_13': ws['F409'].value,
                    'rel_equip_cant_c_m_13': ws['F410'].value,
                    'rel_c_equip_13': rel_c_equip_13,
                    'rel_equip_cant_p_13': ws['G407'].value,
                    'rel_equip_cant_p_b_13': ws['H408'].value,
                    'rel_equip_cant_p_r_13': ws['H409'].value,
                    'rel_equip_cant_p_m_13': ws['H410'].value,
                    'rel_p_equip_13': rel_p_equip_13,
                    'obs_rel_equip_13': ws['I407'].value,
                    'rel_equip_cant_c_14': ws['E412'].value,
                    'rel_equip_cant_c_b_14': ws['F413'].value,
                    'rel_equip_cant_c_r_14': ws['F414'].value,
                    'rel_equip_cant_c_m_14': ws['F415'].value,
                    'rel_c_equip_14': rel_c_equip_14,
                    'rel_equip_cant_p_14': ws['G412'].value,
                    'rel_equip_cant_p_b_14': ws['H413'].value,
                    'rel_equip_cant_p_r_14': ws['H414'].value,
                    'rel_equip_cant_p_m_14': ws['H415'].value,
                    'rel_p_equip_14': rel_p_equip_14,
                    'obs_rel_equip_14': ws['I412'].value,
                    'rel_equip_cant_c_15': ws['E417'].value,
                    'rel_equip_cant_c_b_15': ws['F418'].value,
                    'rel_equip_cant_c_r_15': ws['F419'].value,
                    'rel_equip_cant_c_m_15': ws['F420'].value,
                    'rel_c_equip_15': rel_c_equip_15,
                    'rel_equip_cant_p_15': ws['G417'].value,
                    'rel_equip_cant_p_b_15': ws['H418'].value,
                    'rel_equip_cant_p_r_15': ws['H419'].value,
                    'rel_equip_cant_p_m_15': ws['H420'].value,
                    'rel_p_equip_15': rel_p_equip_15,
                    'obs_rel_equip_15': ws['I417'].value,
                    'rel_equip_cant_c_16': ws['E422'].value,
                    'rel_equip_cant_c_b_16': ws['F423'].value,
                    'rel_equip_cant_c_r_16': ws['F424'].value,
                    'rel_equip_cant_c_m_16': ws['F425'].value,
                    'rel_c_equip_16': rel_c_equip_16,
                    'rel_equip_cant_p_16': ws['G422'].value,
                    'rel_equip_cant_p_b_16': ws['H423'].value,
                    'rel_equip_cant_p_r_16': ws['H424'].value,
                    'rel_equip_cant_p_m_16': ws['H425'].value,
                    'rel_p_equip_16': rel_p_equip_16,
                    'obs_rel_equip_16': ws['I422'].value,
                    'rel_equip_cant_c_17': ws['E427'].value,
                    'rel_equip_cant_c_b_17': ws['F428'].value,
                    'rel_equip_cant_c_r_17': ws['F429'].value,
                    'rel_equip_cant_c_m_17': ws['F430'].value,
                    'rel_c_equip_17': rel_c_equip_17,
                    'rel_equip_cant_p_17': ws['G427'].value,
                    'rel_equip_cant_p_b_17': ws['H428'].value,
                    'rel_equip_cant_p_r_17': ws['H429'].value,
                    'rel_equip_cant_p_m_17': ws['H430'].value,
                    'rel_p_equip_17': rel_p_equip_17,
                    'obs_rel_equip_17': ws['I427'].value,
                    'rel_equip_cant_c_18': ws['E432'].value,
                    'rel_equip_cant_c_b_18': ws['F433'].value,
                    'rel_equip_cant_c_r_18': ws['F434'].value,
                    'rel_equip_cant_c_m_18': ws['F435'].value,
                    'rel_c_equip_18': rel_c_equip_18,
                    'rel_equip_cant_p_18': ws['G432'].value,
                    'rel_equip_cant_p_b_18': ws['H433'].value,
                    'rel_equip_cant_p_r_18': ws['H434'].value,
                    'rel_equip_cant_p_m_18': ws['H435'].value,
                    'rel_p_equip_18': rel_p_equip_18,
                    'obs_rel_equip_18': ws['I432'].value,
                    'rel_equip_cant_c_19': ws['E437'].value,
                    'rel_equip_cant_c_b_19': ws['F438'].value,
                    'rel_equip_cant_c_r_19': ws['F439'].value,
                    'rel_equip_cant_c_m_19': ws['F440'].value,
                    'rel_c_equip_19': rel_c_equip_19,
                    'rel_equip_cant_p_19': ws['G437'].value,
                    'rel_equip_cant_p_b_19': ws['H438'].value,
                    'rel_equip_cant_p_r_19': ws['H439'].value,
                    'rel_equip_cant_p_m_19': ws['H440'].value,
                    'rel_p_equip_19': rel_p_equip_19,
                    'obs_rel_equip_19': ws['I437'].value,
                    'rel_equip_cant_c_20': ws['E442'].value,
                    'rel_equip_cant_c_b_20': ws['F443'].value,
                    'rel_equip_cant_c_r_20': ws['F444'].value,
                    'rel_equip_cant_c_m_20': ws['F445'].value,
                    'rel_c_equip_20': rel_c_equip_20,
                    'rel_equip_cant_p_20': ws['G442'].value,
                    'rel_equip_cant_p_b_20': ws['H443'].value,
                    'rel_equip_cant_p_r_20': ws['H444'].value,
                    'rel_equip_cant_p_m_20': ws['H445'].value,
                    'rel_p_equip_20': rel_p_equip_20,
                    'obs_rel_equip_20': ws['I442'].value,
                    'rel_equip_cant_c_21': ws['E447'].value,
                    'rel_equip_cant_c_b_21': ws['F448'].value,
                    'rel_equip_cant_c_r_21': ws['F449'].value,
                    'rel_equip_cant_c_m_21': ws['F450'].value,
                    'rel_c_equip_21': rel_c_equip_21,
                    'rel_equip_cant_p_21': ws['G447'].value,
                    'rel_equip_cant_p_b_21': ws['H448'].value,
                    'rel_equip_cant_p_r_21': ws['H449'].value,
                    'rel_equip_cant_p_m_21': ws['H450'].value,
                    'rel_p_equip_21': rel_p_equip_21,
                    'obs_rel_equip_21': ws['I447'].value,
                    'rel_equip_cant_c_22': ws['E452'].value,
                    'rel_equip_cant_c_b_22': ws['F453'].value,
                    'rel_equip_cant_c_r_22': ws['F454'].value,
                    'rel_equip_cant_c_m_22': ws['F455'].value,
                    'rel_c_equip_22': rel_c_equip_22,
                    'rel_equip_cant_p_22': ws['G452'].value,
                    'rel_equip_cant_p_b_22': ws['H453'].value,
                    'rel_equip_cant_p_r_22': ws['H454'].value,
                    'rel_equip_cant_p_m_22': ws['H455'].value,
                    'rel_p_equip_22': rel_p_equip_22,
                    'obs_rel_equip_22': ws['I452'].value,
                    'rel_equip_cant_c_23': ws['E457'].value,
                    'rel_equip_cant_c_b_23': ws['F458'].value,
                    'rel_equip_cant_c_r_23': ws['F459'].value,
                    'rel_equip_cant_c_m_23': ws['F460'].value,
                    'rel_c_equip_23': rel_c_equip_23,
                    'rel_equip_cant_p_23': ws['G457'].value,
                    'rel_equip_cant_p_b_23': ws['H458'].value,
                    'rel_equip_cant_p_r_23': ws['H459'].value,
                    'rel_equip_cant_p_m_23': ws['H460'].value,
                    'rel_p_equip_23': rel_p_equip_23,
                    'obs_rel_equip_23': ws['I457'].value,
                    'rel_equip_cant_c_24': ws['E462'].value,
                    'rel_equip_cant_c_b_24': ws['F463'].value,
                    'rel_equip_cant_c_r_24': ws['F464'].value,
                    'rel_equip_cant_c_m_24': ws['F465'].value,
                    'rel_c_equip_24': rel_c_equip_24,
                    'rel_equip_cant_p_24': ws['G462'].value,
                    'rel_equip_cant_p_b_24': ws['H463'].value,
                    'rel_equip_cant_p_r_24': ws['H464'].value,
                    'rel_equip_cant_p_m_24': ws['H465'].value,
                    'rel_p_equip_24': rel_p_equip_24,
                    'obs_rel_equip_24': ws['I462'].value,
                    'rel_equip_cant_c_25': ws['E467'].value,
                    'rel_equip_cant_c_b_25': ws['F468'].value,
                    'rel_equip_cant_c_r_25': ws['F469'].value,
                    'rel_equip_cant_c_m_25': ws['F470'].value,
                    'rel_c_equip_25': rel_c_equip_25,
                    'rel_equip_cant_p_25': ws['G467'].value,
                    'rel_equip_cant_p_b_25': ws['H468'].value,
                    'rel_equip_cant_p_r_25': ws['H469'].value,
                    'rel_equip_cant_p_m_25': ws['H470'].value,
                    'rel_p_equip_25': rel_p_equip_25,
                    'obs_rel_equip_25': ws['I467'].value,
                    'rel_equip_cant_c_26': ws['E472'].value,
                    'rel_equip_cant_c_b_26': ws['F473'].value,
                    'rel_equip_cant_c_r_26': ws['F474'].value,
                    'rel_equip_cant_c_m_26': ws['F475'].value,
                    'rel_c_equip_26': rel_c_equip_26,
                    'rel_equip_cant_p_26': ws['G472'].value,
                    'rel_equip_cant_p_b_26': ws['H473'].value,
                    'rel_equip_cant_p_r_26': ws['H474'].value,
                    'rel_equip_cant_p_m_26': ws['H475'].value,
                    'rel_p_equip_26': rel_p_equip_26,
                    'obs_rel_equip_26': ws['I472'].value,
                    'rel_equip_cant_c_31': ws['E477'].value,
                    'rel_equip_cant_c_b_31': ws['F478'].value,
                    'rel_equip_cant_c_r_31': ws['F479'].value,
                    'rel_equip_cant_c_m_31': ws['F480'].value,
                    'rel_c_equip_31': rel_c_equip_31,
                    'rel_equip_cant_p_31': ws['G477'].value,
                    'rel_equip_cant_p_b_31': ws['H478'].value,
                    'rel_equip_cant_p_r_31': ws['H479'].value,
                    'rel_equip_cant_p_m_31': ws['H480'].value,
                    'rel_p_equip_31': rel_p_equip_31,
                    'obs_rel_equip_31': ws['I477'].value,
                    'rel_equip_cant_c_32': ws['E482'].value,
                    'rel_equip_cant_c_b_32': ws['F483'].value,
                    'rel_equip_cant_c_r_32': ws['F484'].value,
                    'rel_equip_cant_c_m_32': ws['F485'].value,
                    'rel_c_equip_32': rel_c_equip_32,
                    'rel_equip_cant_p_32': ws['G482'].value,
                    'rel_equip_cant_p_b_32': ws['H483'].value,
                    'rel_equip_cant_p_r_32': ws['H484'].value,
                    'rel_equip_cant_p_m_32': ws['H485'].value,
                    'rel_p_equip_32': rel_p_equip_32,
                    'obs_rel_equip_32': ws['I482'].value,
                    'rel_equip_cant_c_33': ws['E487'].value,
                    'rel_equip_cant_c_b_33': ws['F488'].value,
                    'rel_equip_cant_c_r_33': ws['F489'].value,
                    'rel_equip_cant_c_m_33': ws['F490'].value,
                    'rel_c_equip_33': rel_c_equip_33,
                    'rel_equip_cant_p_33': ws['G487'].value,
                    'rel_equip_cant_p_b_33': ws['H488'].value,
                    'rel_equip_cant_p_r_33': ws['H489'].value,
                    'rel_equip_cant_p_m_33': ws['H490'].value,
                    'rel_p_equip_33': rel_p_equip_33,
                    'obs_rel_equip_33': ws['I487'].value,
                    'rel_equip_cant_c_34': ws['E492'].value,
                    'rel_equip_cant_c_b_34': ws['F493'].value,
                    'rel_equip_cant_c_r_34': ws['F494'].value,
                    'rel_equip_cant_c_m_34': ws['F495'].value,
                    'rel_c_equip_34': rel_c_equip_34,
                    'rel_equip_cant_p_34': ws['G492'].value,
                    'rel_equip_cant_p_b_34': ws['H493'].value,
                    'rel_equip_cant_p_r_34': ws['H494'].value,
                    'rel_equip_cant_p_m_34': ws['H495'].value,
                    'rel_p_equip_34': rel_p_equip_34,
                    'obs_rel_equip_34': ws['I492'].value,
                    'rel_equip_cant_c_27': ws['E500'].value,
                    'rel_equip_cant_c_b_27': ws['F501'].value,
                    'rel_equip_cant_c_r_27': ws['F502'].value,
                    'rel_equip_cant_c_m_27': ws['F503'].value,
                    'rel_c_equip_27': rel_c_equip_27,
                    'rel_equip_cant_p_27': ws['G500'].value,
                    'rel_equip_cant_p_b_27': ws['H501'].value,
                    'rel_equip_cant_p_r_27': ws['H502'].value,
                    'rel_equip_cant_p_m_27': ws['H503'].value,
                    'rel_p_equip_27': rel_p_equip_27,
                    'obs_rel_equip_27': ws['I500'].value,
                    'rel_equip_cant_c_28': ws['E505'].value,
                    'rel_equip_cant_c_b_28': ws['F506'].value,
                    'rel_equip_cant_c_r_28': ws['F507'].value,
                    'rel_equip_cant_c_m_28': ws['F508'].value,
                    'rel_c_equip_28': rel_c_equip_28,
                    'rel_equip_cant_p_28': ws['G505'].value,
                    'rel_equip_cant_p_b_28': ws['H506'].value,
                    'rel_equip_cant_p_r_28': ws['H507'].value,
                    'rel_equip_cant_p_m_28': ws['H508'].value,
                    'rel_p_equip_28': rel_p_equip_28,
                    'obs_rel_equip_28': ws['I505'].value,
                    'rel_equip_cant_c_29': ws['E510'].value,
                    'rel_equip_cant_c_b_29': ws['F511'].value,
                    'rel_equip_cant_c_r_29': ws['F512'].value,
                    'rel_equip_cant_c_m_29': ws['F513'].value,
                    'rel_c_equip_29': rel_c_equip_29,
                    'rel_equip_cant_p_29': ws['G510'].value,
                    'rel_equip_cant_p_b_29': ws['H511'].value,
                    'rel_equip_cant_p_r_29': ws['H512'].value,
                    'rel_equip_cant_p_m_29': ws['H513'].value,
                    'rel_p_equip_29': rel_p_equip_29,
                    'obs_rel_equip_29': ws['I510'].value,
                    'rel_equip_cant_c_30': ws['E515'].value,
                    'rel_equip_cant_c_b_30': ws['F516'].value,
                    'rel_equip_cant_c_r_30': ws['F517'].value,
                    'rel_equip_cant_c_m_30': ws['F518'].value,
                    'rel_c_equip_30': rel_c_equip_30,
                    'rel_equip_cant_p_30': ws['G515'].value,
                    'rel_equip_cant_p_b_30': ws['H516'].value,
                    'rel_equip_cant_p_r_30': ws['H517'].value,
                    'rel_equip_cant_p_m_30': ws['H518'].value,
                    'rel_p_equip_30': rel_p_equip_30,
                    'obs_rel_equip_30': ws['I515'].value,
                    't_infra_instal_1': t_infra_instal_1,
                    'esp_infra_instal_1': ws['I523'].value,
                    'obs_infra_instal_1': ws['J523'].value,
                    'p_infra_instal_1': p_infra_instal_1,
                    'g_infra_instal_1': g_infra_instal_1,
                    't_infra_instal_2': t_infra_instal_2,
                    'obs_infra_instal_2': ws['J527'].value,
                    'esp_infra_instal_3': ws['I532'].value,
                    'obs_infra_instal_3': ws['J532'].value,
                    'esp_infra_instal_4': ws['I533'].value,
                    'obs_infra_instal_4': ws['J533'].value,
                    'infra_instal_5': infra_instal_5,
                    'obs_infra_instal_5': ws['J534'].value,
                    'p_infra_instal_5': p_infra_instal_5,
                    'g_infra_instal_5': g_infra_instal_5,
                    'infra_instal_6': infra_instal_6,
                    'obs_infra_instal_6': ws['J535'].value,
                    'p_infra_instal_6': p_infra_instal_6,
                    'g_infra_instal_6': g_infra_instal_6,
                    'infra_instal_7': infra_instal_7,
                    'obs_infra_instal_7': ws['J536'].value,
                    'p_infra_instal_7': p_infra_instal_7,
                    'g_infra_instal_7': g_infra_instal_7,
                    'infra_instal_8': infra_instal_8,
                    'obs_infra_instal_8': ws['J537'].value,
                    'p_infra_instal_8': p_infra_instal_8,
                    'g_infra_instal_8': g_infra_instal_8,
                    'infra_instal_9': infra_instal_9,
                    'obs_infra_instal_9': ws['J538'].value,
                    'p_infra_instal_9': p_infra_instal_9,
                    'g_infra_instal_9': g_infra_instal_9,
                    'infra_instal_10': infra_instal_10,
                    'obs_infra_instal_10': ws['J539'].value,
                    'infra_instal_11': infra_instal_11,
                    'obs_infra_instal_11': ws['J540'].value,
                    'p_infra_instal_11': p_infra_instal_11,
                    'g_infra_instal_11': g_infra_instal_11,
                    'infra_instal_12': infra_instal_12,
                    'obs_infra_instal_12': ws['J541'].value,
                    'p_infra_instal_12': p_infra_instal_12,
                    'g_infra_instal_12': g_infra_instal_12,
                    'infra_instal_13': infra_instal_13,
                    'obs_infra_instal_13': ws['J542'].value,
                    'p_infra_instal_13': p_infra_instal_13,
                    'g_infra_instal_13': g_infra_instal_13,
                    'infra_instal_14': infra_instal_14,
                    'obs_infra_instal_14': ws['J543'].value,
                    'p_infra_instal_14': p_infra_instal_14,
                    'g_infra_instal_14': g_infra_instal_14,
                    'infra_instal_15': infra_instal_15,
                    'obs_infra_instal_15': ws['J544'].value,
                    'p_infra_instal_15': p_infra_instal_15,
                    'g_infra_instal_15': g_infra_instal_15,
                    'infra_instal_16': infra_instal_16,
                    'obs_infra_instal_16': ws['J545'].value,
                    'p_infra_instal_16': p_infra_instal_16,
                    'g_infra_instal_16': g_infra_instal_16,
                    'infra_instal_17': infra_instal_17,
                    'obs_infra_instal_17': ws['J546'].value,
                    'p_infra_instal_17': p_infra_instal_17,
                    'g_infra_instal_17': g_infra_instal_17,
                    'infra_instal_18': infra_instal_18,
                    'obs_infra_instal_18': ws['J547'].value,
                    'p_infra_instal_18': p_infra_instal_18,
                    'g_infra_instal_18': g_infra_instal_18,
                    'infra_instal_19': infra_instal_19,
                    'obs_infra_instal_19': ws['J548'].value,
                    'p_infra_instal_19': p_infra_instal_19,
                    'g_infra_instal_19': g_infra_instal_19,
                    'infra_instal_20': infra_instal_20,
                    'obs_infra_instal_20': ws['J549'].value,
                    'p_infra_instal_20': p_infra_instal_20,
                    'g_infra_instal_20': g_infra_instal_20,
                    'infra_instal_21': infra_instal_21,
                    'obs_infra_instal_21': ws['J550'].value,
                    't_infra_instal_21': t_infra_instal_21,
                    'p_infra_instal_21': p_infra_instal_21,
                    'g_infra_instal_21': g_infra_instal_21,
                    'infra_instal_22': infra_instal_22,
                    'obs_infra_instal_22': ws['J554'].value,
                    'p_infra_instal_22': p_infra_instal_22,
                    'g_infra_instal_22': g_infra_instal_22,
                    'infra_instal_23': infra_instal_23,
                    'obs_infra_instal_23': ws['J555'].value,
                    't_infra_instal_23': t_infra_instal_23,
                    'p_infra_instal_23': p_infra_instal_23,
                    'g_infra_instal_23': g_infra_instal_23,
                    'infra_instal_24': infra_instal_24,
                    'obs_infra_instal_24': ws['J559'].value,
                    'infra_instal_25': infra_instal_25,
                    'obs_infra_instal_25': ws['J560'].value,
                    'infra_instal_26': infra_instal_26,
                    'obs_infra_instal_26': ws['J561'].value,
                    'p_infra_instal_26': p_infra_instal_26,
                    'g_infra_instal_26': g_infra_instal_26,
                    'infra_instal_27': infra_instal_27,
                    'obs_infra_instal_27': ws['J562'].value,
                    'infra_instal_28': infra_instal_28,
                    'obs_infra_instal_28': ws['J563'].value,
                    'bio_s_2': bio_s_2,
                    'obs_bio_s_2': ws['J566'].value,
                    'p_bio_s_2': p_bio_s_2,
                    'g_bio_s_2': g_bio_s_2,
                    'bio_s_3': bio_s_3,
                    'obs_bio_s_3': ws['J567'].value,
                    'bio_s_4': bio_s_4,
                    'obs_bio_s_4': ws['J568'].value,
                    'bio_s_5': bio_s_5,
                    'obs_bio_s_5': ws['J569'].value,
                    'p_bio_s_5': p_bio_s_5,
                    'g_bio_s_5': g_bio_s_5,
                })
            else:
                raise ValidationError('Error al cargar el archivo excel, los datos de validación no coinciden')
            os.unlink(fname)
        else:
            raise ValidationError('Error, la ficha se encuentra en estado Terminado')

    @api.one
    @api.constrains('ficha_supervision')
    def _check_ficha_supervision(self):
        if not (self.txt_ficha_supervision.endswith('.xls') or self.txt_ficha_supervision.endswith('.xlsx')):
            raise ValidationError('Tipo de archivo NO válido')
